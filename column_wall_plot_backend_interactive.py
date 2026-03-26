from __future__ import annotations

import math
import re
import uuid
from copy import deepcopy
from pathlib import Path
from typing import Any

from flask import Flask, jsonify, request, send_file
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

BASE_DIR = Path(__file__).resolve().parent
FRONTEND_FILE = BASE_DIR / "column_wall_plot_frontend_interactive.html"
UPLOAD_DIR = BASE_DIR / "uploads_wall_plotter"
UPLOAD_DIR.mkdir(exist_ok=True)

ALLOWED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}

app = Flask(__name__)
DATASETS: dict[str, dict[str, Any]] = {}
BOUNDARY_DATASETS: dict[str, dict[str, Any]] = {}

REINF_SCHEDULE = [
    {"mark": "A", "bar": 12, "spacing": 300, "area": 754, "t_kn": 320},
    {"mark": "B", "bar": 12, "spacing": 250, "area": 905, "t_kn": 385},
    {"mark": "C", "bar": 12, "spacing": 200, "area": 1131, "t_kn": 481},
    {"mark": "E", "bar": 16, "spacing": 300, "area": 1340, "t_kn": 570},
    {"mark": "D", "bar": 12, "spacing": 150, "area": 1508, "t_kn": 641},
    {"mark": "F", "bar": 16, "spacing": 250, "area": 1608, "t_kn": 684},
    {"mark": "G", "bar": 16, "spacing": 200, "area": 2011, "t_kn": 855},
    {"mark": "J", "bar": 20, "spacing": 300, "area": 2094, "t_kn": 890},
    {"mark": "K", "bar": 20, "spacing": 250, "area": 2513, "t_kn": 1068},
    {"mark": "H", "bar": 16, "spacing": 150, "area": 2681, "t_kn": 1139},
    {"mark": "N", "bar": 24, "spacing": 300, "area": 3016, "t_kn": 1282},
    {"mark": "L", "bar": 20, "spacing": 200, "area": 3142, "t_kn": 1335},
    {"mark": "P", "bar": 24, "spacing": 250, "area": 3619, "t_kn": 1538},
    {"mark": "S", "bar": 28, "spacing": 300, "area": 4105, "t_kn": 1745},
    {"mark": "M", "bar": 20, "spacing": 150, "area": 4189, "t_kn": 1780},
    {"mark": "Q", "bar": 24, "spacing": 200, "area": 4524, "t_kn": 1923},
    {"mark": "T", "bar": 28, "spacing": 250, "area": 4926, "t_kn": 2094},
    {"mark": "W", "bar": 32, "spacing": 300, "area": 5362, "t_kn": 2279},
    {"mark": "R", "bar": 24, "spacing": 150, "area": 6032, "t_kn": 2564},
    {"mark": "U", "bar": 28, "spacing": 200, "area": 6158, "t_kn": 2617},
    {"mark": "X", "bar": 32, "spacing": 250, "area": 6434, "t_kn": 2734},
    {"mark": "AA", "bar": 36, "spacing": 300, "area": 6786, "t_kn": 2884},
    {"mark": "Y", "bar": 32, "spacing": 200, "area": 8042, "t_kn": 3418},
    {"mark": "AB", "bar": 36, "spacing": 250, "area": 8143, "t_kn": 3461},
    {"mark": "V", "bar": 28, "spacing": 150, "area": 8210, "t_kn": 3489},
    {"mark": "AC", "bar": 36, "spacing": 200, "area": 10179, "t_kn": 4326},
    {"mark": "Z", "bar": 32, "spacing": 150, "area": 10723, "t_kn": 4557},
    {"mark": "AD", "bar": 36, "spacing": 150, "area": 13572, "t_kn": 5768},
]

LAP_TABLE_MM = {
    12: {40: 450, 50: 450, 65: 350},
    16: {40: 650, 50: 600, 65: 500},
    20: {40: 850, 50: 800, 65: 650},
    24: {40: 1100, 50: 950, 65: 850},
    28: {40: 1350, 50: 1200, 65: 1050},
    32: {40: 1600, 50: 1450, 65: 1250},
    36: {40: 1900, 50: 1700, 65: 1500},
    40: {40: 2150, 50: 1950, 65: 1750},
}
HOOK_TABLE_MM = {12: 170, 16: 205, 20: 245, 24: 295, 28: 345, 32: 395, 36: 440, 40: 490}
DEFAULT_LIG_BAR_MM = 12
DEFAULT_LIG_COVER_MM = 40
STEEL_DENSITY_FACTOR = 0.00785
WALL_ELEVATION_ALIGNMENT_THRESHOLD = 0.70
GRID_INTERSECTION_TOL_M = 0.03


def norm_name(value: Any) -> str:
    return "".join(ch.lower() for ch in str(value or "") if ch.isalnum())


def text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def as_float(value: Any, default: float | None = None) -> float | None:
    try:
        if value is None or value == "":
            return default
        return float(value)
    except Exception:
        return default


def natural_sort_key(value: Any) -> list[Any]:
    parts = re.split(r"(\d+)", str(value or ""))
    out: list[Any] = []
    for part in parts:
        out.append(int(part) if part.isdigit() else part.lower())
    return out




def truthy(value: Any) -> bool:
    return text(value).strip().lower() in {"1", "true", "t", "yes", "y", "on"}


def bar_area_mm2(bar: int | None) -> float | None:
    if not bar or int(bar) <= 0:
        return None
    return math.pi * float(bar) * float(bar) / 4.0


def bar_mass_per_m_kg(bar: int | None) -> float:
    area = bar_area_mm2(bar)
    return float(STEEL_DENSITY_FACTOR * area) if area else 0.0


def ratio_from_value(value: Any, fc_mpa: float | None) -> float | None:
    raw = as_float(value)
    if raw is None:
        return None
    raw = float(raw)
    if abs(raw) <= 1.0:
        return raw
    if fc_mpa and float(fc_mpa) > 0.0:
        return raw / float(fc_mpa)
    return None


def first_not_none(*values: Any) -> Any:
    for value in values:
        if value is not None:
            return value
    return None

def find_sheet(workbook, target: str, required: bool = True):
    target_n = norm_name(target)
    for ws in workbook.worksheets:
        title_n = norm_name(ws.title)
        if title_n == target_n or target_n in title_n:
            return ws
    if required:
        raise KeyError(f"Required sheet not found: {target}")
    return None


def read_table(ws) -> tuple[list[str], list[str], list[dict[str, Any]]]:
    headers = [text(c.value) for c in ws[2]]
    units = [text(c.value) for c in ws[3]]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row is None:
            continue
        if all(v is None for v in row[: len(headers)]):
            continue
        rows.append({headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]})
    return headers, units, rows


def build_story_elevations(
    stories_top_down: list[str],
    height_by_story: dict[str, float | None],
    base_story_name: str | None = None,
    base_story_elev_m: float | None = None,
) -> dict[str, float]:
    if not stories_top_down:
        return {}
    bottom_up = list(reversed(stories_top_down))
    if base_story_name and base_story_name not in bottom_up:
        bottom_up.append(base_story_name)
    start_story = bottom_up[0]
    elev = {start_story: float(base_story_elev_m or 0.0) if start_story == base_story_name else 0.0}
    for i in range(1, len(bottom_up)):
        elev[bottom_up[i]] = elev[bottom_up[i - 1]] + float(height_by_story.get(bottom_up[i]) or 0.0)
    return elev


def read_base_story_definition(workbook) -> tuple[str | None, float | None]:
    ws = find_sheet(workbook, "Tower and Base Story Definition", required=False)
    if ws is None:
        return None, None
    try:
        _, _, rows = read_table(ws)
    except Exception:
        return None, None
    for row in rows:
        name = text(row.get("BSName") or row.get("Base Story") or row.get("Base Story Name"))
        elev = as_float(row.get("BSElev") or row.get("Base Elevation") or row.get("Base Elev"))
        if name:
            return name, elev
    return None, None


def rotated_rect_corners(cx: float, cy: float, width: float, thickness: float, angle_deg: float) -> list[list[float]]:
    half_w = width / 2.0
    half_t = thickness / 2.0
    ang = math.radians(angle_deg)
    ux, uy = math.cos(ang), math.sin(ang)
    vx, vy = -math.sin(ang), math.cos(ang)
    return [
        [cx - half_w * ux - half_t * vx, cy - half_w * uy - half_t * vy],
        [cx + half_w * ux - half_t * vx, cy + half_w * uy - half_t * vy],
        [cx + half_w * ux + half_t * vx, cy + half_w * uy + half_t * vy],
        [cx - half_w * ux + half_t * vx, cy - half_w * uy + half_t * vy],
    ]


def axis_from_grid_type(value: str) -> str | None:
    v = norm_name(value)
    if "x" in v or "vertical" in v:
        return "X"
    if "y" in v or "horizontal" in v:
        return "Y"
    return None


def element_intersects_grid(axis: str, ordinate_m: float, corners: list[list[float]]) -> bool:
    if not corners:
        return False
    idx = 0 if axis == "X" else 1
    vals = [p[idx] for p in corners]
    return (min(vals) - GRID_INTERSECTION_TOL_M) <= ordinate_m <= (max(vals) + GRID_INTERSECTION_TOL_M)


def alignment_to_grid(axis: str, angle_deg: float) -> float:
    ang = math.radians(angle_deg)
    ux = abs(math.cos(ang))
    uy = abs(math.sin(ang))
    return uy if axis == "X" else ux


def projected_station_and_width(axis: str, item: dict[str, Any]) -> tuple[float, float]:
    coords = [p[1] for p in item["corners"]] if axis == "X" else [p[0] for p in item["corners"]]
    return (sum(coords) / len(coords), max(max(coords) - min(coords), 0.05))


def parse_reinf_string(value: Any) -> tuple[int | None, int | None]:
    m = re.match(r"^\s*N\s*(\d+)\s*-\s*(\d+)\s*$", text(value), re.I)
    if not m:
        return None, None
    return int(m.group(1)), int(m.group(2))


def area_from_reinf(value: Any) -> float | None:
    bar, spacing = parse_reinf_string(value)
    if not bar or not spacing or spacing <= 0:
        return None
    return 2.0 * (math.pi * bar * bar / 4.0) * (1000.0 / spacing)


def parse_fc_mpa(material: Any) -> float | None:
    m = re.search(r"(\d+(?:\.\d+)?)", text(material))
    return float(m.group(1)) if m else None


def lap_length_mm(bar: int | None, fc_mpa: float | None) -> float:
    if not bar or bar not in LAP_TABLE_MM:
        return 0.0
    fc = float(fc_mpa or 40.0)
    bucket = 65 if fc >= 65 else 50 if fc >= 50 else 40
    return float(LAP_TABLE_MM[bar][bucket])


def hook_length_mm(bar: int | None) -> float:
    return float(HOOK_TABLE_MM.get(int(bar or 0), 0.0))


def bars_per_layer(span_m: float | None, spacing_mm: int | None) -> int | None:
    span_mm = max(float(span_m or 0.0) * 1000.0, 0.0)
    if span_mm <= 0.0 or not spacing_mm or float(spacing_mm) <= 0.0:
        return None
    return int(math.ceil(span_mm / float(spacing_mm)) + 1)


def choose_schedule(required_mm2pm: float | None) -> dict[str, Any] | None:
    if required_mm2pm is None or required_mm2pm <= 0:
        return None
    for item in REINF_SCHEDULE:
        if item["area"] >= required_mm2pm:
            return deepcopy(item)
    return deepcopy(REINF_SCHEDULE[-1])


def schedule_text(required_mm2pm: float | None) -> str:
    item = choose_schedule(required_mm2pm)
    return f"N{item['bar']}-{item['spacing']}" if item else "—"


def current_reinf_value(override_value: str | None, fallback: str | None) -> str:
    return text(override_value) or text(fallback) or "—"


def reinforcement_strings(base: dict[str, Any], override: dict[str, Any] | None) -> tuple[str, str, str]:
    vertical = current_reinf_value((override or {}).get("vertical"), base.get("seed_vertical"))
    horizontal = current_reinf_value((override or {}).get("horizontal"), base.get("seed_horizontal"))
    return vertical, horizontal, f"V: {vertical} / H: {horizontal}"


def rate_kgpm3_for_wall(vertical: str | None, horizontal: str | None, thickness_m: float | None, length_m: float | None, story_height_m: float | None, fc_mpa: float | None) -> float | None:
    t = float(thickness_m or 0.0)
    L = float(length_m or 0.0)
    H = float(story_height_m or 0.0)
    if t <= 0 or L <= 0 or H <= 0:
        return None
    total = 0.0
    v_bar, v_spacing = parse_reinf_string(vertical)
    v_count = bars_per_layer(L, v_spacing)
    if v_bar and v_count:
        total += 2.0 * float(v_count) * (H + lap_length_mm(v_bar, fc_mpa) / 1000.0) * bar_mass_per_m_kg(v_bar)
    h_bar, h_spacing = parse_reinf_string(horizontal)
    h_count = bars_per_layer(H, h_spacing)
    if h_bar and h_count:
        u_len_m = max(float(thickness_m or 0.0) * 1000.0 - 2.0 * DEFAULT_LIG_COVER_MM + 2.4 * lap_length_mm(h_bar, fc_mpa), 0.0) / 1000.0
        total += 2.0 * float(h_count) * L * bar_mass_per_m_kg(h_bar)
        total += 2.0 * float(h_count) * u_len_m * bar_mass_per_m_kg(h_bar)
    volume = t * L * H
    return (total / volume) if total > 0 and volume > 0 else None


def make_item(kind: str, label: str, source_story: str, plot_story: str, material: str, angle_deg: float, cgx_m: float, cgy_m: float, width_m: float, thickness_m: float, depth_m: float = 0.0) -> dict[str, Any]:
    return {
        "key": f"{kind}|{source_story}|{label}",
        "kind": kind,
        "label": label,
        "source_story": source_story,
        "plot_story": plot_story,
        "material": material or "",
        "angle_deg": float(angle_deg or 0.0),
        "cgx_m": float(cgx_m or 0.0),
        "cgy_m": float(cgy_m or 0.0),
        "width_m": max(float(width_m or 0.0), 0.0),
        "thickness_m": max(float(thickness_m or 0.0), 0.0),
        "depth_m": max(float(depth_m or 0.0), 0.0),
        "length_m": max(float(width_m or 0.0), 0.0),
        "corners": rotated_rect_corners(float(cgx_m or 0.0), float(cgy_m or 0.0), max(float(width_m or 0.0), 0.0), max(float(thickness_m or 0.0), 0.0), float(angle_deg or 0.0)),
    }



def parse_pier_design_sheet(ws, walls: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    _, _, rows = read_table(ws)
    wall_lookup = {(w["source_story"], w["label"]): w for w in walls}
    grouped: dict[str, dict[str, Any]] = {}
    for row in rows:
        story = text(row.get("Story"))
        label = text(row.get("Pier Label") or row.get("Pier"))
        if not story or not label:
            continue
        wall = wall_lookup.get((story, label))
        thickness_mm = as_float(row.get("Thickness"))
        if thickness_mm is None and wall:
            thickness_mm = float((wall or {}).get("thickness_m") or 0.0) * 1000.0
        req_pct = as_float(row.get("Required Reinf. Percentage"))
        cur_pct = as_float(row.get("Current Reinf. Percentage"))
        shear = as_float(row.get("Shear Rebar"))

        required_v = ((req_pct / 100.0) * thickness_mm * 1000.0) if (req_pct is not None and thickness_mm is not None) else None
        current_v = ((cur_pct / 100.0) * thickness_mm * 1000.0) if (cur_pct is not None and thickness_mm is not None) else None
        required_h = shear if shear is not None else None

        key = f"wall|{story}|{label}"
        entry = grouped.setdefault(key, {})
        if required_v is not None:
            entry["required_vertical_mm2pm"] = max(float(entry.get("required_vertical_mm2pm") or 0.0), float(required_v))
        if current_v is not None:
            entry["current_vertical_mm2pm"] = max(float(entry.get("current_vertical_mm2pm") or 0.0), float(current_v))
        if required_h is not None:
            entry["required_horizontal_mm2pm"] = max(float(entry.get("required_horizontal_mm2pm") or 0.0), float(required_h))
    for entry in grouped.values():
        entry["suggested_vertical"] = schedule_text(entry.get("required_vertical_mm2pm"))
        entry["suggested_horizontal"] = schedule_text(entry.get("required_horizontal_mm2pm"))
        entry["seed_vertical"] = schedule_text(entry.get("current_vertical_mm2pm") or entry.get("required_vertical_mm2pm"))
        entry["seed_horizontal"] = schedule_text(entry.get("required_horizontal_mm2pm"))
    return grouped


def find_first_header(headers: list[str], patterns: list[str]) -> str | None:
    normalized = {header: norm_name(header) for header in headers if text(header)}
    for pattern in patterns:
        pattern_n = norm_name(pattern)
        for header, header_n in normalized.items():
            if header_n == pattern_n or pattern_n in header_n:
                return header
    return None


def parse_boundary_workbook(path: Path) -> dict[str, dict[str, Any]]:
    wb = load_workbook(path, data_only=True, read_only=False)
    ws = find_sheet(wb, "Pier Dgn Sum")
    headers, _, rows = read_table(ws)
    story_header = find_first_header(headers, ["Story", "Pier Story"])
    pier_header = find_first_header(headers, ["Pier Label", "Pier", "Pier Name"])
    left_header = find_first_header(headers, [
        "Boundary Zone Left", "Boundary Left", "Boundary Length Left", "Left Boundary Length",
        "Boundary Zone Length Left", "Length of Left Boundary", "Left Boundary Zone", "Left Boundary",
        "Boundary Element Left", "Left Boundary Element", "BZ Left", "Left BZ", "BE Left", "Left BE",
    ])
    right_header = find_first_header(headers, [
        "Boundary Zone Right", "Boundary Right", "Boundary Length Right", "Right Boundary Length",
        "Boundary Zone Length Right", "Length of Right Boundary", "Right Boundary Zone", "Right Boundary",
        "Boundary Element Right", "Right Boundary Element", "BZ Right", "Right BZ", "BE Right", "Right BE",
    ])
    general_stress_header = find_first_header(headers, [
        "Compressive Stress", "Compression Stress", "Comp Stress", "Max Compressive Stress",
        "Wall Compressive Stress", "Boundary Compressive Stress", "Stress/FC", "Stress Ratio"
    ])
    left_stress_header = find_first_header(headers, [
        "Left Compressive Stress", "Compressive Stress Left", "Left Compression Stress", "Compression Stress Left",
        "Left Comp Stress", "Comp Stress Left", "Boundary Compressive Stress Left", "Left Boundary Stress",
        "BE Stress Left", "Stress Left"
    ])
    right_stress_header = find_first_header(headers, [
        "Right Compressive Stress", "Compressive Stress Right", "Right Compression Stress", "Compression Stress Right",
        "Right Comp Stress", "Comp Stress Right", "Boundary Compressive Stress Right", "Right Boundary Stress",
        "BE Stress Right", "Stress Right"
    ])
    general_limit_header = find_first_header(headers, [
        "Compression Limit", "Compressive Limit", "Boundary Compression Limit", "Stress Limit", "Limit/FC", "Limit Ratio"
    ])
    left_limit_header = find_first_header(headers, [
        "Left Compression Limit", "Compression Limit Left", "Left Compressive Limit", "Compressive Limit Left",
        "Left Stress Limit", "Stress Limit Left", "Boundary Compression Limit Left", "BE Limit Left", "Limit Left"
    ])
    right_limit_header = find_first_header(headers, [
        "Right Compression Limit", "Compression Limit Right", "Right Compressive Limit", "Compressive Limit Right",
        "Right Stress Limit", "Stress Limit Right", "Boundary Compression Limit Right", "BE Limit Right", "Limit Right"
    ])
    if not story_header or not pier_header or (not left_header and not right_header):
        raise KeyError("Could not identify boundary-zone columns in 'Pier Dgn Sum'.")
    out: dict[str, dict[str, Any]] = {}
    for row in rows:
        story = text(row.get(story_header))
        label = text(row.get(pier_header))
        if not story or not label:
            continue
        key = f"wall|{story}|{label}"
        entry = out.setdefault(key, {})
        left_mm = max(float(as_float(row.get(left_header), 0.0) or 0.0), 0.0) if left_header else 0.0
        right_mm = max(float(as_float(row.get(right_header), 0.0) or 0.0), 0.0) if right_header else 0.0
        entry["left_mm"] = left_mm
        entry["right_mm"] = right_mm
        entry["stress_raw"] = as_float(row.get(general_stress_header)) if general_stress_header else None
        entry["left_stress_raw"] = as_float(row.get(left_stress_header)) if left_stress_header else None
        entry["right_stress_raw"] = as_float(row.get(right_stress_header)) if right_stress_header else None
        entry["limit_raw"] = as_float(row.get(general_limit_header)) if general_limit_header else None
        entry["left_limit_raw"] = as_float(row.get(left_limit_header)) if left_limit_header else None
        entry["right_limit_raw"] = as_float(row.get(right_limit_header)) if right_limit_header else None
        if left_mm <= 0.0 and right_mm <= 0.0 and all(entry.get(k) is None for k in ("stress_raw", "left_stress_raw", "right_stress_raw", "limit_raw", "left_limit_raw", "right_limit_raw")):
            out.pop(key, None)
    return out


def build_ligature_data(item: dict[str, Any], dataset: dict[str, Any]) -> dict[str, Any]:
    boundary = deepcopy(dataset.get("boundary_by_key", {}).get(item.get("key"), {}))
    left_mm = max(float(boundary.get("left_mm") or 0.0), 0.0)
    right_mm = max(float(boundary.get("right_mm") or 0.0), 0.0)
    fc_mpa = parse_fc_mpa(item.get("material"))
    general_stress_ratio = ratio_from_value(boundary.get("stress_raw"), fc_mpa)
    left_stress_ratio = first_not_none(ratio_from_value(boundary.get("left_stress_raw"), fc_mpa), general_stress_ratio)
    right_stress_ratio = first_not_none(ratio_from_value(boundary.get("right_stress_raw"), fc_mpa), general_stress_ratio)
    general_limit_ratio = ratio_from_value(boundary.get("limit_raw"), fc_mpa)
    left_limit_ratio = first_not_none(ratio_from_value(boundary.get("left_limit_raw"), fc_mpa), general_limit_ratio)
    right_limit_ratio = first_not_none(ratio_from_value(boundary.get("right_limit_raw"), fc_mpa), general_limit_ratio)
    has_boundary = left_mm > 0.0 or right_mm > 0.0
    _, vertical_spacing = parse_reinf_string(item.get("current_vertical"))
    _, horizontal_spacing = parse_reinf_string(item.get("current_horizontal"))
    base = {
        "has_boundary": has_boundary,
        "boundary_left_mm": left_mm,
        "boundary_right_mm": right_mm,
        "boundary_stress_ratio": general_stress_ratio,
        "boundary_left_stress_ratio": left_stress_ratio,
        "boundary_right_stress_ratio": right_stress_ratio,
        "boundary_limit_ratio": general_limit_ratio,
        "boundary_left_limit_ratio": left_limit_ratio,
        "boundary_right_limit_ratio": right_limit_ratio,
        "horizontal_spacing_mm": horizontal_spacing,
        "horizontal_spacing_warning": bool(has_boundary and horizontal_spacing and float(horizontal_spacing) > 200.0),
    }
    if not has_boundary:
        return {**base, "ligs": "—", "ligs_n": None, "ligs_sets": None, "ligs_length_mm": None}
    if not vertical_spacing or vertical_spacing <= 0 or not horizontal_spacing or horizontal_spacing <= 0:
        return {**base, "ligs": "—", "ligs_n": None, "ligs_sets": None, "ligs_length_mm": None}
    ligs_n = sum(2 * (math.ceil(length_mm / float(vertical_spacing)) + 1) for length_mm in (left_mm, right_mm) if length_mm > 0.0)
    story_height_mm = max(float(dataset.get("height_by_story", {}).get(item.get("source_story")) or 0.0) * 1000.0, 0.0)
    set_spacing_mm = min(float(horizontal_spacing), 200.0)
    ligs_sets = math.ceil(story_height_mm / set_spacing_mm) if story_height_mm > 0.0 and set_spacing_mm > 0.0 else None
    wall_thickness_mm = max(float(item.get("thickness_m") or 0.0) * 1000.0, 0.0)
    ligs_length_mm = max(wall_thickness_mm - 2.0 * DEFAULT_LIG_COVER_MM + 2.0 * hook_length_mm(DEFAULT_LIG_BAR_MM), 0.0) if wall_thickness_mm > 0.0 else None
    ligs_text = f"{ligs_n}xN{DEFAULT_LIG_BAR_MM}"
    if ligs_sets is not None:
        ligs_text += f" ({ligs_sets} sets)"
    return {**base, "ligs": ligs_text, "ligs_n": ligs_n, "ligs_sets": ligs_sets, "ligs_length_mm": ligs_length_mm}


def boundary_hatch_style(fc_mpa: float | None, stress_ratio: float | None) -> str | None:
    fc = float(fc_mpa or 0.0)
    if stress_ratio is None:
        return "diag-red"
    if stress_ratio > 0.2:
        return "diag-blue"
    if 0.15 < stress_ratio < 0.2:
        return "dot-red" if fc <= 50.0 else "dot-blue"
    return "diag-red"


def body_hatch_style(fc_mpa: float | None) -> str | None:
    fc = float(fc_mpa or 0.0)
    if fc > 65.0:
        return "diag-blue"
    if fc > 50.0:
        return "dot-blue"
    return None


def wall_quantities(item: dict[str, Any], dataset: dict[str, Any]) -> dict[str, float | int | None]:
    if item.get("kind") != "wall":
        return {
            "steel_kg": None,
            "concrete_volume_m3": None,
            "rate_kgpm3": None,
            "vertical_kg": None,
            "horizontal_kg": None,
            "ligs_kg": None,
            "vertical_bar_count_per_layer": None,
            "horizontal_bar_count_per_layer": None,
        }
    L = max(float(item.get("width_m") or 0.0), 0.0)
    t = max(float(item.get("thickness_m") or 0.0), 0.0)
    H = max(float(dataset.get("height_by_story", {}).get(item.get("source_story")) or 0.0), 0.0)
    volume = L * t * H if L > 0 and t > 0 and H > 0 else None
    fc = parse_fc_mpa(item.get("material"))

    vertical_bar, vertical_spacing = parse_reinf_string(item.get("current_vertical"))
    horizontal_bar, horizontal_spacing = parse_reinf_string(item.get("current_horizontal"))
    vertical_bar_count_per_layer = bars_per_layer(L, vertical_spacing)
    horizontal_bar_count_per_layer = bars_per_layer(H, horizontal_spacing)

    vertical_kg = None
    if vertical_bar and vertical_bar_count_per_layer and H > 0:
        vertical_len_m = H + lap_length_mm(vertical_bar, fc) / 1000.0
        vertical_kg = 2.0 * float(vertical_bar_count_per_layer) * vertical_len_m * bar_mass_per_m_kg(vertical_bar)

    horizontal_kg = None
    if horizontal_bar and horizontal_bar_count_per_layer and L > 0:
        u_bar_length_m = max(t * 1000.0 - 2.0 * DEFAULT_LIG_COVER_MM + 2.4 * lap_length_mm(horizontal_bar, fc), 0.0) / 1000.0
        straight_kg = 2.0 * float(horizontal_bar_count_per_layer) * L * bar_mass_per_m_kg(horizontal_bar)
        u_bar_kg = 2.0 * float(horizontal_bar_count_per_layer) * u_bar_length_m * bar_mass_per_m_kg(horizontal_bar)
        horizontal_kg = straight_kg + u_bar_kg

    ligs_kg = None
    ligs_n = item.get("ligs_n")
    ligs_sets = item.get("ligs_sets")
    ligs_length_mm = item.get("ligs_length_mm")
    if ligs_n and ligs_sets and ligs_length_mm:
        ligs_kg = float(ligs_n) * float(ligs_sets) * (float(ligs_length_mm) / 1000.0) * bar_mass_per_m_kg(DEFAULT_LIG_BAR_MM)

    steel_components = [value for value in (vertical_kg, horizontal_kg, ligs_kg) if value is not None]
    steel_kg = sum(steel_components) if steel_components else None
    rate = (steel_kg / volume) if (steel_kg is not None and volume and volume > 0) else None
    return {
        "steel_kg": steel_kg,
        "concrete_volume_m3": volume,
        "rate_kgpm3": rate,
        "vertical_kg": vertical_kg,
        "horizontal_kg": horizontal_kg,
        "ligs_kg": ligs_kg,
        "vertical_bar_count_per_layer": vertical_bar_count_per_layer,
        "horizontal_bar_count_per_layer": horizontal_bar_count_per_layer,
    }


def parse_spandrel_design_sheet(ws) -> dict[str, dict[str, Any]]:
    _, _, rows = read_table(ws)
    grouped: dict[str, dict[str, Any]] = {}
    for row in rows:
        story = text(row.get("Story"))
        label = text(row.get("Spandrel"))
        if not story or not label:
            continue
        key = f"spandrel|{story}|{label}"
        entry = grouped.setdefault(key, {})
        av_vert = as_float(row.get("Av Vert"))
        av_horz = as_float(row.get("Av Horz"))
        if av_vert is not None:
            entry["required_vertical_mm2pm"] = max(float(entry.get("required_vertical_mm2pm") or 0.0), av_vert)
        if av_horz is not None:
            entry["required_horizontal_mm2pm"] = max(float(entry.get("required_horizontal_mm2pm") or 0.0), av_horz)
    for entry in grouped.values():
        entry["suggested_vertical"] = schedule_text(entry.get("required_vertical_mm2pm"))
        entry["suggested_horizontal"] = schedule_text(entry.get("required_horizontal_mm2pm"))
        entry["seed_vertical"] = entry["suggested_vertical"]
        entry["seed_horizontal"] = entry["suggested_horizontal"]
    return grouped


def parse_spandrels(rows: list[dict[str, Any]], below_story_by_source: dict[str, str]) -> list[dict[str, Any]]:
    out = []
    for row in rows:
        source_story = text(row.get("Story"))
        label = text(row.get("Spandrel"))
        plot_story = below_story_by_source.get(source_story)
        if not source_story or not label or not plot_story:
            continue
        lx, ly = as_float(row.get("CG Left X")), as_float(row.get("CG Left Y"))
        rx, ry = as_float(row.get("CG Right X")), as_float(row.get("CG Right Y"))
        if None in (lx, ly, rx, ry):
            continue
        dx, dy = float(rx) - float(lx), float(ry) - float(ly)
        endpoint_len = math.hypot(dx, dy)
        length_m = endpoint_len if endpoint_len > 1e-6 else (as_float(row.get("Length"), 0.0) or 0.0) / 1000.0
        if length_m <= 0:
            continue
        angle_deg = math.degrees(math.atan2(dy, dx)) if endpoint_len > 1e-6 else 0.0
        thickness_m = max(((as_float(row.get("Thickness Left"), 0.0) or 0.0) + (as_float(row.get("Thickness Right"), 0.0) or 0.0)) / 2000.0, (as_float(row.get("Thickness Left"), 0.0) or 0.0) / 1000.0, (as_float(row.get("Thickness Right"), 0.0) or 0.0) / 1000.0)
        depth_m = max(((as_float(row.get("Depth Left"), 0.0) or 0.0) + (as_float(row.get("Depth Right"), 0.0) or 0.0)) / 2000.0, (as_float(row.get("Depth Left"), 0.0) or 0.0) / 1000.0, (as_float(row.get("Depth Right"), 0.0) or 0.0) / 1000.0)
        item = make_item("spandrel", label, source_story, plot_story, text(row.get("Material")), angle_deg, (float(lx) + float(rx)) / 2.0, (float(ly) + float(ry)) / 2.0, length_m, thickness_m, depth_m)
        item["line_x1_m"], item["line_y1_m"], item["line_x2_m"], item["line_y2_m"] = float(lx), float(ly), float(rx), float(ry)
        out.append(item)
    return out


def build_ductility_profile(dataset: dict[str, Any], ductility_level: str | int | None) -> dict[str, Any]:
    level = str(ductility_level or "1")
    profile: dict[str, Any] = {"active": level in {"2", "3"}, "ductility_level": level, "rows": [], "min_pct_by_plot_story": {}, "fc_max_mpa": None, "lw_max_m": None, "base_pct": None, "two_lw_height_m": None, "two_storeys_height_m": None, "governing_height_m": None, "governing_condition": None, "base_story_count": 0}
    if level not in {"2", "3"}:
        return profile
    fc_values = [parse_fc_mpa(w.get("material")) for w in dataset.get("walls", [])]
    fc_values = [v for v in fc_values if v is not None]
    lw_max = max((float(w.get("width_m") or 0.0) for w in dataset.get("walls", [])), default=0.0)
    # Preserve structural level order from the workbook: base story first, then levels above.
    plot_stories = list(dataset.get("available_plot_stories", []))
    if not fc_values or lw_max <= 0 or not plot_stories:
        return profile
    fc_max = max(fc_values)
    base_pct = max(0.25, 100.0 * 0.7 * math.sqrt(fc_max) / 500.0)
    two_lw = 2.0 * lw_max
    two_storeys = 0.0
    for plot_story in plot_stories[:2]:
        two_storeys += float(dataset["height_by_story"].get(dataset["source_story_by_plot"].get(plot_story)) or 0.0)
    governing_h = max(two_lw, two_storeys)
    governing_condition = "2×Lw" if two_lw >= two_storeys else "2 storeys"
    cum_h, base_story_count = 0.0, 0
    for plot_story in plot_stories:
        cum_h += float(dataset["height_by_story"].get(dataset["source_story_by_plot"].get(plot_story)) or 0.0)
        base_story_count += 1
        if cum_h >= governing_h - 1e-9:
            break
    base_story_count = max(base_story_count, 2 if len(plot_stories) >= 2 else 1)
    rows, min_by = [], {}
    for idx, plot_story in enumerate(plot_stories):
        pct = max(0.25, base_pct * (0.9 ** idx))
        rows.append({"plot_story": plot_story, "min_pct": pct})
        min_by[plot_story] = pct
        if pct <= 0.2500001:
            break
    profile.update({"rows": rows, "min_pct_by_plot_story": min_by, "fc_max_mpa": fc_max, "lw_max_m": lw_max, "base_pct": base_pct, "two_lw_height_m": two_lw, "two_storeys_height_m": two_storeys, "governing_height_m": governing_h, "governing_condition": governing_condition, "base_story_count": base_story_count})
    return profile


def parse_workbook(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True, read_only=False)
    story_ws = find_sheet(wb, "Story Definitions")
    pier_ws = find_sheet(wb, "Pier Section Properties")
    grid_ws = find_sheet(wb, "Grid Definitions - Grid Lines", required=False)
    spandrel_ws = find_sheet(wb, "Spandrel Section Properties", required=False)
    pier_dgn_ws = find_sheet(wb, "Pier Dgn Sum", required=False)
    span_dgn_ws = find_sheet(wb, "Span Dgn Sum", required=False)
    _, _, story_rows = read_table(story_ws)
    _, _, pier_rows = read_table(pier_ws)
    _, _, grid_rows = read_table(grid_ws) if grid_ws else ([], [], [])
    _, _, spandrel_rows = read_table(spandrel_ws) if spandrel_ws else ([], [], [])
    base_story_name, base_story_elev_m = read_base_story_definition(wb)
    stories_top_down = [text(r.get("Name")) for r in story_rows if text(r.get("Name"))]
    if base_story_name and base_story_name not in stories_top_down:
        stories_top_down = stories_top_down + [base_story_name]
    height_by_story = {text(r.get("Name")): as_float(r.get("Height")) for r in story_rows if text(r.get("Name"))}
    if base_story_name and base_story_name not in height_by_story:
        height_by_story[base_story_name] = 0.0
    story_elevation_m = build_story_elevations(stories_top_down, height_by_story, base_story_name, base_story_elev_m)
    below_story_by_source = {stories_top_down[i]: stories_top_down[i + 1] for i in range(len(stories_top_down) - 1)}
    source_story_by_plot = {v: k for k, v in below_story_by_source.items()}
    walls = []
    for row in pier_rows:
        source_story = text(row.get("Story"))
        label = text(row.get("Pier"))
        plot_story = below_story_by_source.get(source_story)
        if not source_story or not label or not plot_story:
            continue
        walls.append(make_item("wall", label, source_story, plot_story, text(row.get("Material")), as_float(row.get("AxisAngle"), 0.0) or 0.0, as_float(row.get("CG Bottom X"), 0.0) or 0.0, as_float(row.get("CG Bottom Y"), 0.0) or 0.0, (as_float(row.get("Width Bottom"), 0.0) or 0.0) / 1000.0, (as_float(row.get("Thickness Bottom"), 0.0) or 0.0) / 1000.0, 0.0))
    spandrels = parse_spandrels(spandrel_rows, below_story_by_source) if spandrel_rows else []
    design_by_key: dict[str, dict[str, Any]] = {}
    if pier_dgn_ws:
        design_by_key.update(parse_pier_design_sheet(pier_dgn_ws, walls))
    if span_dgn_ws:
        design_by_key.update(parse_spandrel_design_sheet(span_dgn_ws))
    systems: dict[str, list[dict[str, Any]]] = {}
    for row in grid_rows:
        system = text(row.get("Name"))
        if not system or text(row.get("Visible") or "Yes").lower() == "no":
            continue
        axis = axis_from_grid_type(text(row.get("Grid Line Type")))
        ordinate = as_float(row.get("Ordinate"))
        grid_id = text(row.get("ID"))
        if not axis or ordinate is None or not grid_id:
            continue
        systems.setdefault(system, []).append({"key": f"{system}|{axis}|{grid_id}", "system": system, "id": grid_id, "axis": axis, "ordinate_m": float(ordinate), "label": grid_id})
    core_grids = systems.get("CORE", [])
    core_grids.sort(key=lambda g: natural_sort_key(g["id"]))
    # Plot levels must follow structural order from the base upward, not alphanumeric order.
    stories_bottom_up = list(reversed(stories_top_down))
    available_plot_stories = [story for story in stories_bottom_up if any(el["plot_story"] == story for el in (walls + spandrels))]
    dataset = {"filename": path.name, "stories_top_down": stories_top_down, "available_plot_stories": available_plot_stories, "grouping_story_list": stories_bottom_up, "height_by_story": height_by_story, "story_elevation_m": story_elevation_m, "source_story_by_plot": source_story_by_plot, "base_story_name": base_story_name, "base_story_elev_m": base_story_elev_m, "grids": core_grids, "default_grid_key": core_grids[0]["key"] if core_grids else None, "walls": walls, "spandrels": spandrels, "design_by_key": design_by_key, "overrides": {}, "has_pier_design": bool(pier_dgn_ws), "has_spandrels": bool(spandrels), "boundary_workbook": None, "boundary_by_key": {}}
    dataset["ductility_profile"] = build_ductility_profile(dataset, "2")
    return dataset


def attach_design_and_overrides(dataset: dict[str, Any], item: dict[str, Any], include_boundary: bool = True) -> dict[str, Any]:
    result = deepcopy(item)
    design = deepcopy(dataset.get("design_by_key", {}).get(item["key"], {}))
    override = deepcopy(dataset.get("overrides", {}).get(item["key"], {}))
    fc = parse_fc_mpa(item.get("material"))
    result["fc_mpa"] = fc
    min_pct = None
    min_vertical = None
    if item["kind"] == "wall":
        min_pct = dataset.get("ductility_profile", {}).get("min_pct_by_plot_story", {}).get(item["plot_story"])
        if min_pct is not None:
            min_vertical = (min_pct / 100.0) * float(item.get("thickness_m") or 0.0) * 1_000_000.0
    required_v = design.get("required_vertical_mm2pm")
    if min_vertical is not None:
        required_v = max(float(required_v or 0.0), float(min_vertical))
    required_h = design.get("required_horizontal_mm2pm")
    seed_v = schedule_text(required_v)
    seed_h = schedule_text(required_h)
    result.update({"required_vertical_mm2pm": required_v, "required_horizontal_mm2pm": required_h, "minimum_vertical_mm2pm": min_vertical, "minimum_vertical_pct": min_pct, "suggested_vertical": schedule_text(required_v), "suggested_horizontal": schedule_text(required_h), "seed_vertical": seed_v, "seed_horizontal": seed_h, "override_vertical": text(override.get("vertical")), "override_horizontal": text(override.get("horizontal"))})
    v, h, combo = reinforcement_strings(result, override)
    result["display_vertical"] = v
    result["display_horizontal"] = h
    result["display_reinforcement"] = combo
    result["current_vertical"] = v
    result["current_horizontal"] = h
    result["current_vertical_mm2pm"] = area_from_reinf(v)
    result["current_horizontal_mm2pm"] = area_from_reinf(h)
    result["is_overridden"] = bool(result.get("override_vertical") or result.get("override_horizontal"))
    result["is_vertical_ok"] = result["required_vertical_mm2pm"] is None or (result["current_vertical_mm2pm"] or 0.0) >= float(result["required_vertical_mm2pm"] or 0.0) - 1e-9
    result["is_horizontal_ok"] = result["required_horizontal_mm2pm"] is None or (result["current_horizontal_mm2pm"] or 0.0) >= float(result["required_horizontal_mm2pm"] or 0.0) - 1e-9
    if result["kind"] == "wall":
        if include_boundary:
            result.update(build_ligature_data(result, dataset))
        else:
            result.update({
                "ligs": "—", "ligs_n": None, "ligs_sets": None, "ligs_length_mm": None,
                "has_boundary": False, "boundary_left_mm": 0.0, "boundary_right_mm": 0.0,
                "boundary_stress_ratio": None, "boundary_left_stress_ratio": None, "boundary_right_stress_ratio": None,
                "boundary_limit_ratio": None, "boundary_left_limit_ratio": None, "boundary_right_limit_ratio": None,
                "horizontal_spacing_mm": None, "horizontal_spacing_warning": False,
            })
        result["body_hatch_style"] = body_hatch_style(fc) if include_boundary else None
        result["boundary_left_hatch_style"] = boundary_hatch_style(fc, result.get("boundary_left_stress_ratio")) if include_boundary and result.get("boundary_left_mm", 0) > 0 else None
        result["boundary_right_hatch_style"] = boundary_hatch_style(fc, result.get("boundary_right_stress_ratio")) if include_boundary and result.get("boundary_right_mm", 0) > 0 else None
        qty = wall_quantities(result, dataset)
        result["steel_kg"] = qty["steel_kg"]
        result["concrete_volume_m3"] = qty["concrete_volume_m3"]
        result["current_rate_kgpm3"] = qty["rate_kgpm3"]
        result["vertical_kg"] = qty["vertical_kg"]
        result["horizontal_kg"] = qty["horizontal_kg"]
        result["ligs_kg"] = qty["ligs_kg"]
        result["vertical_bar_count_per_layer"] = qty["vertical_bar_count_per_layer"]
        result["horizontal_bar_count_per_layer"] = qty["horizontal_bar_count_per_layer"]
    else:
        result.update({"ligs": "—", "ligs_n": None, "ligs_sets": None, "ligs_length_mm": None, "has_boundary": False, "boundary_left_mm": 0.0, "boundary_right_mm": 0.0, "body_hatch_style": None, "boundary_left_hatch_style": None, "boundary_right_hatch_style": None, "steel_kg": None, "concrete_volume_m3": None, "current_rate_kgpm3": None, "vertical_kg": None, "horizontal_kg": None, "ligs_kg": None, "vertical_bar_count_per_layer": None, "horizontal_bar_count_per_layer": None})
    return result


def build_reinforcement_rows(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for item in items:
        if item["kind"] != "wall":
            continue
        rows.append({
            "key": item["key"], "label": item["label"], "story": item["plot_story"], "material": item.get("material") or "",
            "thickness_m": item.get("thickness_m"), "vertical": item.get("current_vertical") or "—", "horizontal": item.get("current_horizontal") or "—",
            "ligs": item.get("ligs") or "—", "ligs_n": item.get("ligs_n"), "ligs_sets": item.get("ligs_sets"), "ligs_length_mm": item.get("ligs_length_mm"),
            "vertical_area": item.get("current_vertical_mm2pm"), "horizontal_area": item.get("current_horizontal_mm2pm"), "rate_kgpm3": item.get("current_rate_kgpm3"),
            "steel_kg": item.get("steel_kg"), "vertical_kg": item.get("vertical_kg"), "horizontal_kg": item.get("horizontal_kg"), "ligs_kg": item.get("ligs_kg"), "concrete_volume_m3": item.get("concrete_volume_m3"),
            "vertical_bar_count_per_layer": item.get("vertical_bar_count_per_layer"), "horizontal_bar_count_per_layer": item.get("horizontal_bar_count_per_layer"),
            "required_vertical_area": item.get("required_vertical_mm2pm"), "required_horizontal_area": item.get("required_horizontal_mm2pm"),
            "vertical_ok": bool(item.get("is_vertical_ok")), "horizontal_ok": bool(item.get("is_horizontal_ok")), "overridden": bool(item.get("is_overridden")),
            "has_boundary": bool(item.get("has_boundary")), "horizontal_spacing_mm": item.get("horizontal_spacing_mm"), "horizontal_spacing_warning": bool(item.get("horizontal_spacing_warning")),
        })
    rows.sort(key=lambda r: natural_sort_key(r["label"]))
    return rows


def build_all_reinforcement_rows(dataset: dict[str, Any], include_boundary: bool = True) -> list[dict[str, Any]]:
    return build_reinforcement_rows([attach_design_and_overrides(dataset, w, include_boundary=include_boundary) for w in dataset.get("walls", [])])


def build_floor_summary_rows(dataset: dict[str, Any], include_boundary: bool = True) -> list[dict[str, Any]]:
    rows = []
    for plot_story in dataset.get("available_plot_stories", []):
        walls = [attach_design_and_overrides(dataset, w, include_boundary=include_boundary) for w in dataset.get("walls", []) if w["plot_story"] == plot_story]
        spandrels = [attach_design_and_overrides(dataset, s, include_boundary=include_boundary) for s in dataset.get("spandrels", []) if s["plot_story"] == plot_story]
        total_steel_kg = sum(float(w.get("steel_kg") or 0.0) for w in walls)
        total_volume_m3 = sum(float(w.get("concrete_volume_m3") or 0.0) for w in walls)
        rows.append({"plot_story": plot_story, "wall_count": len(walls), "spandrel_count": len(spandrels), "total_steel_kg": total_steel_kg, "total_volume_m3": total_volume_m3, "avg_rate_kgpm3": (total_steel_kg / total_volume_m3) if total_volume_m3 > 0 else None})
    return rows


def build_plan_payload(dataset: dict[str, Any], story_name: str) -> dict[str, Any]:
    include_boundary = truthy(request.args.get("show_boundary"))
    items = [attach_design_and_overrides(dataset, el, include_boundary=include_boundary) for el in (dataset["walls"] + dataset["spandrels"]) if el["plot_story"] == story_name]
    items.sort(key=lambda d: (d["kind"], natural_sort_key(d["label"])))
    return {"ok": True, "mode": "plan", "filename": dataset["filename"], "story": story_name, "summary": {"mode": "plan", "plot_story": story_name, "wall_count": sum(1 for i in items if i["kind"] == "wall"), "spandrel_count": sum(1 for i in items if i["kind"] == "spandrel")}, "items": items, "has_boundary_workbook": bool(dataset.get("boundary_workbook"))}


def build_elevation_payload(dataset: dict[str, Any], grid_key: str) -> dict[str, Any]:
    grid = next((g for g in dataset.get("grids", []) if g["key"] == grid_key), None)
    if not grid:
        raise KeyError(f"Grid '{grid_key}' not found.")
    include_boundary = truthy(request.args.get("show_boundary"))
    items = []
    for base in dataset["walls"] + dataset["spandrels"]:
        if not element_intersects_grid(grid["axis"], float(grid["ordinate_m"]), base["corners"]):
            continue
        if alignment_to_grid(grid["axis"], float(base["angle_deg"])) < WALL_ELEVATION_ALIGNMENT_THRESHOLD:
            continue
        item = attach_design_and_overrides(dataset, base, include_boundary=include_boundary)
        station_m, display_width_m = projected_station_and_width(grid["axis"], item)
        if item["kind"] == "wall":
            z0_m = float(dataset["story_elevation_m"].get(item["plot_story"], 0.0))
            z1_m = z0_m + float(dataset["height_by_story"].get(item["source_story"]) or 0.0)
            x0_m = station_m - display_width_m / 2.0
            x1_m = station_m + display_width_m / 2.0
        else:
            if grid["axis"] == "X":
                x0_m = min(float(item.get("line_y1_m") or item["cgy_m"]), float(item.get("line_y2_m") or item["cgy_m"]))
                x1_m = max(float(item.get("line_y1_m") or item["cgy_m"]), float(item.get("line_y2_m") or item["cgy_m"]))
            else:
                x0_m = min(float(item.get("line_x1_m") or item["cgx_m"]), float(item.get("line_x2_m") or item["cgx_m"]))
                x1_m = max(float(item.get("line_x1_m") or item["cgx_m"]), float(item.get("line_x2_m") or item["cgx_m"]))
            if abs(x1_m - x0_m) < 0.03:
                x0_m -= float(item.get("thickness_m") or 0.0) / 2.0
                x1_m += float(item.get("thickness_m") or 0.0) / 2.0
            z1_m = float(dataset["story_elevation_m"].get(item["source_story"], 0.0))
            z0_m = z1_m - float(item.get("depth_m") or 0.0)
            station_m = (x0_m + x1_m) / 2.0
            display_width_m = x1_m - x0_m
        items.append({**item, "grid_key": grid["key"], "grid_id": grid["id"], "grid_axis": grid["axis"], "grid_ordinate_m": grid["ordinate_m"], "station_m": station_m, "display_width_m": display_width_m, "x0_m": x0_m, "x1_m": x1_m, "z0_m": z0_m, "z1_m": z1_m})
    items.sort(key=lambda d: (d["z0_m"], d["station_m"], d["kind"], natural_sort_key(d["label"])))
    xs = [v for item in items for v in (float(item["x0_m"]), float(item["x1_m"]))]
    zs = [v for item in items for v in (float(item["z0_m"]), float(item["z1_m"]))]
    return {"ok": True, "mode": "elevation", "filename": dataset["filename"], "grid_key": grid_key, "summary": {"mode": "elevation", "grid_key": grid["key"], "grid_id": grid["id"], "grid_axis": grid["axis"], "wall_count": sum(1 for i in items if i["kind"] == "wall"), "spandrel_count": sum(1 for i in items if i["kind"] == "spandrel"), "x_min_m": min(xs) if xs else None, "x_max_m": max(xs) if xs else None, "z_min_m": min(zs) if zs else 0.0, "z_max_m": max(zs) if zs else None}, "items": items, "story_elevation_m": dataset["story_elevation_m"], "stories_top_down": dataset["stories_top_down"], "has_boundary_workbook": bool(dataset.get("boundary_workbook"))}


def remove_boundary_dataset(boundary_info: dict[str, Any] | None) -> None:
    if not boundary_info:
        return
    path = boundary_info.get("path")
    if path:
        Path(path).unlink(missing_ok=True)
    boundary_id = text(boundary_info.get("id"))
    if boundary_id:
        BOUNDARY_DATASETS.pop(boundary_id, None)


def clear_boundary_from_dataset(dataset: dict[str, Any]) -> None:
    remove_boundary_dataset(dataset.get("boundary_workbook"))
    dataset["boundary_workbook"] = None
    dataset["boundary_by_key"] = {}


def remove_dataset(dataset_id: str) -> None:
    dataset = DATASETS.pop(dataset_id, None)
    if not dataset:
        return
    clear_boundary_from_dataset(dataset)
    path = dataset.get("path")
    if path:
        Path(path).unlink(missing_ok=True)


@app.errorhandler(404)
def handle_404(err):
    if request.path.startswith("/api/"):
        return jsonify({"ok": False, "error": f"API route not found: {request.path}"}), 404
    return err


@app.errorhandler(405)
def handle_405(err):
    if request.path.startswith("/api/"):
        return jsonify({"ok": False, "error": f"Method not allowed for API route: {request.path}"}), 405
    return err


@app.errorhandler(500)
def handle_500(err):
    if request.path.startswith("/api/"):
        return jsonify({"ok": False, "error": "Internal server error."}), 500
    return err


@app.get("/")
def root():
    return send_file(FRONTEND_FILE)


@app.post("/api/upload")
def upload_workbook():
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "No file uploaded."}), 400
    file = request.files["file"]
    if not file or not file.filename:
        return jsonify({"ok": False, "error": "No file selected."}), 400
    filename = secure_filename(file.filename)
    ext = Path(filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({"ok": False, "error": "Please upload an Excel workbook (.xlsx, .xlsm, .xltx, .xltm)."}), 400
    previous_dataset_id = text(request.form.get("previous_dataset_id"))
    dataset_id = uuid.uuid4().hex[:12]
    save_path = UPLOAD_DIR / f"{dataset_id}_{filename}"
    file.save(save_path)
    try:
        parsed = parse_workbook(save_path)
    except Exception as exc:
        save_path.unlink(missing_ok=True)
        return jsonify({"ok": False, "error": f"Could not read workbook: {exc}"}), 400
    parsed["path"] = str(save_path)
    DATASETS[dataset_id] = parsed
    if previous_dataset_id and previous_dataset_id != dataset_id:
        try:
            remove_dataset(previous_dataset_id)
        except Exception:
            pass
    return jsonify({"ok": True, "dataset_id": dataset_id, "filename": parsed["filename"], "stories": parsed["available_plot_stories"], "default_story": parsed["available_plot_stories"][0] if parsed["available_plot_stories"] else None, "grids": parsed["grids"], "default_grid_key": parsed["default_grid_key"], "has_pier_design": parsed["has_pier_design"], "has_spandrels": parsed["has_spandrels"], "reinforcement_schedule": REINF_SCHEDULE, "ductility_profile": parsed.get("ductility_profile"), "grouping_story_list": parsed.get("grouping_story_list", []) or parsed.get("available_plot_stories", []), "has_boundary_workbook": False})


@app.post("/api/upload-boundary")
def upload_boundary_workbook():
    dataset_id = text(request.form.get("dataset_id"))
    dataset = DATASETS.get(dataset_id)
    if not dataset:
        return jsonify({"ok": False, "error": "Main workbook dataset not found."}), 404
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "No file uploaded."}), 400
    file = request.files["file"]
    if not file or not file.filename:
        return jsonify({"ok": False, "error": "No file selected."}), 400
    filename = secure_filename(file.filename)
    ext = Path(filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({"ok": False, "error": "Please upload an Excel workbook (.xlsx, .xlsm, .xltx, .xltm)."}), 400
    clear_boundary_from_dataset(dataset)
    boundary_id = uuid.uuid4().hex[:12]
    save_path = UPLOAD_DIR / f"boundary_{dataset_id}_{boundary_id}_{filename}"
    file.save(save_path)
    try:
        boundary_by_key = parse_boundary_workbook(save_path)
    except Exception as exc:
        save_path.unlink(missing_ok=True)
        return jsonify({"ok": False, "error": f"Could not read boundary workbook: {exc}"}), 400
    boundary_info = {"id": boundary_id, "filename": filename, "path": str(save_path)}
    BOUNDARY_DATASETS[boundary_id] = boundary_info
    dataset["boundary_workbook"] = boundary_info
    dataset["boundary_by_key"] = boundary_by_key
    return jsonify({"ok": True, "dataset_id": dataset_id, "boundary_dataset_id": boundary_id, "filename": filename, "boundary_count": len(boundary_by_key)})


@app.post("/api/dataset/<dataset_id>/boundary/clear")
def clear_boundary_workbook(dataset_id: str):
    dataset = DATASETS.get(dataset_id)
    if not dataset:
        return jsonify({"ok": False, "error": "Dataset not found."}), 404
    clear_boundary_from_dataset(dataset)
    return jsonify({"ok": True})


@app.post("/api/dataset/<dataset_id>/ductility")
def set_ductility(dataset_id: str):
    dataset = DATASETS.get(dataset_id)
    if not dataset:
        return jsonify({"ok": False, "error": "Dataset not found."}), 404
    payload = request.get_json(silent=True) or {}
    dataset["ductility_profile"] = build_ductility_profile(dataset, payload.get("level"))
    return jsonify({"ok": True, "ductility_profile": dataset["ductility_profile"], "grouping_story_list": dataset.get("grouping_story_list", []) or dataset.get("available_plot_stories", [])})


@app.get("/api/dataset/<dataset_id>/plan/<story_name>")
def get_plan(dataset_id: str, story_name: str):
    dataset = DATASETS.get(dataset_id)
    if not dataset:
        return jsonify({"ok": False, "error": "Dataset not found."}), 404
    if story_name not in dataset["available_plot_stories"]:
        return jsonify({"ok": False, "error": f"Story '{story_name}' not found."}), 404
    return jsonify(build_plan_payload(dataset, story_name))


@app.get("/api/dataset/<dataset_id>/elevation/<path:grid_key>")
def get_elevation(dataset_id: str, grid_key: str):
    dataset = DATASETS.get(dataset_id)
    if not dataset:
        return jsonify({"ok": False, "error": "Dataset not found."}), 404
    try:
        return jsonify(build_elevation_payload(dataset, grid_key))
    except KeyError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 404


@app.get("/api/dataset/<dataset_id>/reinforcement")
def get_reinforcement(dataset_id: str):
    dataset = DATASETS.get(dataset_id)
    if not dataset:
        return jsonify({"ok": False, "error": "Dataset not found."}), 404
    include_boundary = truthy(request.args.get("use_boundary"))
    return jsonify({"ok": True, "rows": build_all_reinforcement_rows(dataset, include_boundary=include_boundary)})


@app.get("/api/dataset/<dataset_id>/summary")
def get_summary(dataset_id: str):
    dataset = DATASETS.get(dataset_id)
    if not dataset:
        return jsonify({"ok": False, "error": "Dataset not found."}), 404
    include_boundary = truthy(request.args.get("use_boundary"))
    return jsonify({"ok": True, "rows": build_floor_summary_rows(dataset, include_boundary=include_boundary), "ductility_profile": dataset.get("ductility_profile"), "grouping_story_list": dataset.get("grouping_story_list", []) or dataset.get("available_plot_stories", [])})


@app.get("/api/dataset/<dataset_id>/item/<path:item_key>")
def get_item(dataset_id: str, item_key: str):
    dataset = DATASETS.get(dataset_id)
    if not dataset:
        return jsonify({"ok": False, "error": "Dataset not found."}), 404
    base = next((el for el in (dataset["walls"] + dataset["spandrels"]) if el["key"] == item_key), None)
    if not base:
        return jsonify({"ok": False, "error": "Element not found."}), 404
    return jsonify({"ok": True, "item": attach_design_and_overrides(dataset, base, include_boundary=bool(dataset.get("boundary_workbook")))})


@app.post("/api/dataset/<dataset_id>/override")
def set_override(dataset_id: str):
    dataset = DATASETS.get(dataset_id)
    if not dataset:
        return jsonify({"ok": False, "error": "Dataset not found."}), 404
    payload = request.get_json(silent=True) or {}
    key = text(payload.get("key"))
    if not key:
        return jsonify({"ok": False, "error": "Missing element key."}), 400
    vertical = text(payload.get("vertical"))
    horizontal = text(payload.get("horizontal"))
    if not vertical and not horizontal:
        dataset["overrides"].pop(key, None)
    else:
        dataset["overrides"][key] = {"vertical": vertical, "horizontal": horizontal}
    base = next((el for el in (dataset["walls"] + dataset["spandrels"]) if el["key"] == key), None)
    item = attach_design_and_overrides(dataset, base, include_boundary=bool(dataset.get("boundary_workbook"))) if base else None
    warnings = []
    if item:
        if not item.get("is_vertical_ok", True):
            warnings.append("Vertical override area is below the required minimum.")
        if not item.get("is_horizontal_ok", True):
            warnings.append("Horizontal override area is below the required minimum.")
    return jsonify({"ok": True, "item": item, "warning": " ".join(warnings)})


@app.post("/api/dataset/<dataset_id>/override/reset-all")
def reset_all_overrides(dataset_id: str):
    dataset = DATASETS.get(dataset_id)
    if not dataset:
        return jsonify({"ok": False, "error": "Dataset not found."}), 404
    dataset["overrides"] = {}
    return jsonify({"ok": True})

@app.get("/api/health")
def health():
    return jsonify({"ok": True})


if __name__ == "__main__":
    app.run(debug=False, use_reloader=False, host="127.0.0.1", port=5000)
