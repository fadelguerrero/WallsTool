
from __future__ import annotations

import math
import re
import uuid
from copy import deepcopy
from pathlib import Path
from typing import Any

from flask import Flask, jsonify, request, send_file
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

BASE_DIR = Path(__file__).resolve().parent
FRONTEND_FILE = BASE_DIR / "htk_walls_frontend.html"
UPLOAD_DIR = BASE_DIR / "uploads_wall_plotter"
UPLOAD_DIR.mkdir(exist_ok=True)

ALLOWED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}

app = Flask(__name__)
DATASETS: dict[str, dict[str, Any]] = {}
BOUNDARY_DATASETS: dict[str, dict[str, Any]] = {}

DEFAULT_WALL_REINF_SCHEDULE = [{'mark': 'A',
  'class': 'N',
  'bar': 12,
  'spacing': 300,
  'layers': 2,
  'area': 753.9822368615504,
  'use': True,
  'lig_bar': 10,
  'lig_spacing1': 180,
  'lig_spacing2': 90,
  'dowel_size': 24,
  'dowel_spacing': 600,
  'dowel_area': 753.9822368615504},
 {'mark': 'B',
  'class': 'N',
  'bar': 12,
  'spacing': 250,
  'layers': 2,
  'area': 904.7786842338604,
  'use': True,
  'lig_bar': 10,
  'lig_spacing1': 180,
  'lig_spacing2': 90,
  'dowel_size': 24,
  'dowel_spacing': 500,
  'dowel_area': 904.7786842338604},
 {'mark': 'C',
  'class': 'N',
  'bar': 12,
  'spacing': 200,
  'layers': 2,
  'area': 1130.9733552923256,
  'use': True,
  'lig_bar': 10,
  'lig_spacing1': 180,
  'lig_spacing2': 90,
  'dowel_size': 24,
  'dowel_spacing': 400,
  'dowel_area': 1130.9733552923256},
 {'mark': 'E',
  'class': 'N',
  'bar': 16,
  'spacing': 300,
  'layers': 2,
  'area': 1340.412865531645,
  'use': True,
  'lig_bar': 10,
  'lig_spacing1': 240,
  'lig_spacing2': 120,
  'dowel_size': 32,
  'dowel_spacing': 600,
  'dowel_area': 1340.412865531645},
 {'mark': 'D',
  'class': 'N',
  'bar': 12,
  'spacing': 150,
  'layers': 2,
  'area': 0.0,
  'use': False,
  'lig_bar': 10,
  'lig_spacing1': 180,
  'lig_spacing2': 90,
  'dowel_size': 24,
  'dowel_spacing': 0,
  'dowel_area': 0.0},
 {'mark': 'F',
  'class': 'N',
  'bar': 16,
  'spacing': 250,
  'layers': 2,
  'area': 1608.495438637974,
  'use': True,
  'lig_bar': 10,
  'lig_spacing1': 240,
  'lig_spacing2': 120,
  'dowel_size': 32,
  'dowel_spacing': 500,
  'dowel_area': 1608.495438637974},
 {'mark': 'G',
  'class': 'N',
  'bar': 16,
  'spacing': 200,
  'layers': 2,
  'area': 2010.6192982974674,
  'use': True,
  'lig_bar': 10,
  'lig_spacing1': 240,
  'lig_spacing2': 120,
  'dowel_size': 32,
  'dowel_spacing': 400,
  'dowel_area': 2010.6192982974674},
 {'mark': 'J',
  'class': 'N',
  'bar': 20,
  'spacing': 300,
  'layers': 2,
  'area': 0.0,
  'use': False,
  'lig_bar': 10,
  'lig_spacing1': 300,
  'lig_spacing2': 150,
  'dowel_size': 36,
  'dowel_spacing': 0,
  'dowel_area': 0.0},
 {'mark': 'K',
  'class': 'N',
  'bar': 20,
  'spacing': 250,
  'layers': 2,
  'area': 2513.2741228718346,
  'use': True,
  'lig_bar': 10,
  'lig_spacing1': 300,
  'lig_spacing2': 150,
  'dowel_size': 36,
  'dowel_spacing': 400,
  'dowel_area': 2544.6900494077327},
 {'mark': 'H',
  'class': 'N',
  'bar': 16,
  'spacing': 150,
  'layers': 2,
  'area': 2680.82573106329,
  'use': True,
  'lig_bar': 10,
  'lig_spacing1': 240,
  'lig_spacing2': 120,
  'dowel_size': 32,
  'dowel_spacing': 300,
  'dowel_area': 2680.82573106329},
 {'mark': 'N',
  'class': 'N',
  'bar': 24,
  'spacing': 300,
  'layers': 2,
  'area': 0.0,
  'use': False,
  'lig_bar': 10,
  'lig_spacing1': 300,
  'lig_spacing2': 180,
  'dowel_size': 36,
  'dowel_spacing': 0,
  'dowel_area': 0.0},
 {'mark': 'L',
  'class': 'N',
  'bar': 20,
  'spacing': 200,
  'layers': 2,
  'area': 3141.5926535897934,
  'use': True,
  'lig_bar': 10,
  'lig_spacing1': 300,
  'lig_spacing2': 150,
  'dowel_size': 36,
  'dowel_spacing': 300,
  'dowel_area': 3392.9200658769764},
 {'mark': 'P',
  'class': 'N',
  'bar': 24,
  'spacing': 250,
  'layers': 2,
  'area': 3619.1147369354417,
  'use': True,
  'lig_bar': 10,
  'lig_spacing1': 300,
  'lig_spacing2': 180,
  'dowel_size': 36,
  'dowel_spacing': 275,
  'dowel_area': 3701.3673445930654},
 {'mark': 'S',
  'class': 'N',
  'bar': 28,
  'spacing': 300,
  'layers': 2,
  'area': 0.0,
  'use': False,
  'lig_bar': 10,
  'lig_spacing1': 300,
  'lig_spacing2': 200,
  'dowel_size': 36,
  'dowel_spacing': 0,
  'dowel_area': 0.0},
 {'mark': 'M',
  'class': 'N',
  'bar': 20,
  'spacing': 150,
  'layers': 2,
  'area': 4188.790204786391,
  'use': True,
  'lig_bar': 10,
  'lig_spacing1': 300,
  'lig_spacing2': 150,
  'dowel_size': 36,
  'dowel_spacing': 225,
  'dowel_area': 4523.893421169302},
 {'mark': 'Q',
  'class': 'N',
  'bar': 24,
  'spacing': 200,
  'layers': 2,
  'area': 4523.893421169302,
  'use': True,
  'lig_bar': 10,
  'lig_spacing1': 300,
  'lig_spacing2': 180,
  'dowel_size': 36,
  'dowel_spacing': 225,
  'dowel_area': 4523.893421169302},
 {'mark': 'T',
  'class': 'N',
  'bar': 28,
  'spacing': 250,
  'layers': 2,
  'area': 4926.017280828795,
  'use': True,
  'lig_bar': 10,
  'lig_spacing1': 300,
  'lig_spacing2': 200,
  'dowel_size': 36,
  'dowel_spacing': 200,
  'dowel_area': 5089.380098815465},
 {'mark': 'W',
  'class': 'N',
  'bar': 32,
  'spacing': 300,
  'layers': 2,
  'area': 5361.65146212658,
  'use': True,
  'lig_bar': 12,
  'lig_spacing1': 300,
  'lig_spacing2': 250,
  'dowel_size': 0,
  'dowel_spacing': 0,
  'dowel_area': 0.0},
 {'mark': 'R',
  'class': 'N',
  'bar': 24,
  'spacing': 150,
  'layers': 2,
  'area': 6031.857894892403,
  'use': True,
  'lig_bar': 10,
  'lig_spacing1': 300,
  'lig_spacing2': 180,
  'dowel_size': 36,
  'dowel_spacing': 0,
  'dowel_area': 0.0},
 {'mark': 'U',
  'class': 'N',
  'bar': 28,
  'spacing': 200,
  'layers': 2,
  'area': 6157.5216010359945,
  'use': True,
  'lig_bar': 10,
  'lig_spacing1': 300,
  'lig_spacing2': 200,
  'dowel_size': 36,
  'dowel_spacing': 0,
  'dowel_area': 0.0},
 {'mark': 'X',
  'class': 'N',
  'bar': 32,
  'spacing': 250,
  'layers': 2,
  'area': 6433.981754551896,
  'use': True,
  'lig_bar': 12,
  'lig_spacing1': 300,
  'lig_spacing2': 250,
  'dowel_size': 0,
  'dowel_spacing': 0,
  'dowel_area': 0.0}]
DEFAULT_SPANDREL_LONG_SCHEDULE = [{'mark': 'A', 'class': 'N', 'bar': 20, 'bars': 2, 'area': 0.0, 'use': False},
 {'mark': 'B', 'class': 'N', 'bar': 24, 'bars': 2, 'area': 904.7786842338604, 'use': True},
 {'mark': 'C', 'class': 'N', 'bar': 28, 'bars': 2, 'area': 1231.5043202071988, 'use': True},
 {'mark': 'D', 'class': 'N', 'bar': 32, 'bars': 2, 'area': 1608.495438637974, 'use': True},
 {'mark': 'E', 'class': 'N', 'bar': 36, 'bars': 2, 'area': 2035.7520395261859, 'use': True},
 {'mark': 'F', 'class': 'N', 'bar': 28, 'bars': 4, 'area': 2463.0086404143976, 'use': True},
 {'mark': 'G', 'class': 'N', 'bar': 32, 'bars': 4, 'area': 3216.990877275948, 'use': True},
 {'mark': 'J', 'class': 'N', 'bar': 36, 'bars': 4, 'area': 4071.5040790523717, 'use': True},
 {'mark': 'K', 'class': 'N', 'bar': 32, 'bars': 6, 'area': 4825.486315913922, 'use': True},
 {'mark': 'I', 'class': 'N', 'bar': 36, 'bars': 6, 'area': 6107.256118578558, 'use': True},
 {'mark': 'L', 'class': 'N', 'bar': 40, 'bars': 6, 'area': 7539.822368615503, 'use': True},
 {'mark': 'M', 'class': 'N', 'bar': 40, 'bars': 9, 'area': 11309.733552923255, 'use': True},
 {'mark': 'N', 'class': 'N', 'bar': 40, 'bars': 12, 'area': 15079.644737231007, 'use': True},
 {'mark': 'O', 'class': 'N', 'bar': 40, 'bars': 15, 'area': 0.0, 'use': False},
 {'mark': 'P', 'class': 'N', 'bar': 40, 'bars': 18, 'area': 0.0, 'use': False}]
DEFAULT_SPANDREL_LIG_SCHEDULE = [{'mark': 'L1', 'class': 'N', 'bar': 12, 'legs': 2, 'spacing': 200, 'area': 1130.9733552923256, 'use': True},
 {'mark': 'L2', 'class': 'N', 'bar': 12, 'legs': 2, 'spacing': 150, 'area': 1507.9644737231008, 'use': True},
 {'mark': 'L3', 'class': 'N', 'bar': 16, 'legs': 2, 'spacing': 200, 'area': 2010.6192982974674, 'use': True},
 {'mark': 'L4', 'class': 'N', 'bar': 16, 'legs': 2, 'spacing': 150, 'area': 2680.82573106329, 'use': True},
 {'mark': 'L5', 'class': 'N', 'bar': 16, 'legs': 2, 'spacing': 125, 'area': 3216.990877275948, 'use': True},
 {'mark': 'L6', 'class': 'N', 'bar': 16, 'legs': 4, 'spacing': 175, 'area': 4595.7012532513545, 'use': True},
 {'mark': 'L7', 'class': 'N', 'bar': 16, 'legs': 4, 'spacing': 150, 'area': 5361.65146212658, 'use': True},
 {'mark': 'L8', 'class': 'N', 'bar': 16, 'legs': 4, 'spacing': 125, 'area': 6433.981754551896, 'use': True},
 {'mark': 'L9', 'class': 'N', 'bar': 20, 'legs': 4, 'spacing': 175, 'area': 7180.783208205242, 'use': True},
 {'mark': 'L10', 'class': 'N', 'bar': 20, 'legs': 4, 'spacing': 150, 'area': 8377.580409572782, 'use': True},
 {'mark': 'L11', 'class': 'N', 'bar': 20, 'legs': 4, 'spacing': 125, 'area': 10053.096491487338, 'use': True},
 {'mark': 'L13', 'class': 'N', 'bar': 24, 'legs': 4, 'spacing': 125, 'area': 14476.458947741767, 'use': True},
 {'mark': 'L15', 'class': 'N', 'bar': 24, 'legs': 6, 'spacing': 125, 'area': 21714.68842161265, 'use': True},
 {'mark': 'L16', 'class': 'N', 'bar': 24, 'legs': 8, 'spacing': 125, 'area': 28952.917895483533, 'use': True},
 {'mark': 'L17', 'class': 'N', 'bar': 24, 'legs': 10, 'spacing': 125, 'area': 36191.14736935442, 'use': True}]
DEFAULT_WALL_LAP_TABLE_MM = {12: 450, 16: 650, 20: 850, 24: 1000, 28: 1300, 32: 1550, 36: 1850, 40: 2150}
DEFAULT_WALL_DOWEL_LAP_TABLE_MM = {24: 800, 32: 1200, 36: 1500}
DEFAULT_SPANDREL_LAP_TABLE_MM = {12: {'lap': 450, 'cog': 170},
 16: {'lap': 650, 'cog': 200},
 20: {'lap': 850, 'cog': 240},
 24: {'lap': 1000, 'cog': 280},
 28: {'lap': 1300, 'cog': 330},
 32: {'lap': 1550, 'cog': 370},
 36: {'lap': 1850, 'cog': 420},
 40: {'lap': 2150, 'cog': 490}}
DEFAULT_WALL_HORIZONTAL_SPLICE_FACTORS = [(1.0, 0.5), (3.0, 1.0), (9.0, 2.0), (15.0, 3.0), (18.0, 4.0)]

DEFAULT_COVER_MM = 20.0
DEFAULT_DUCTILITY_LEVEL = "2"
DEFAULT_WALL_TYPE = "INSITU"
HOOK_TABLE_MM = {12: 170, 16: 205, 20: 245, 24: 295, 28: 345, 32: 395, 36: 440, 40: 490}
STEEL_DENSITY_FACTOR = 7850.0
WALL_ELEVATION_ALIGNMENT_THRESHOLD = 0.70
GRID_INTERSECTION_TOL_M = 0.03


def norm_name(value: Any) -> str:
    return "".join(ch.lower() for ch in str(value or "") if ch.isalnum())


def text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def as_float(value: Any, default: float | None = None) -> float | None:
    try:
        if value is None or value == "":
            return default
        return float(value)
    except Exception:
        return default


def natural_sort_key(value: Any) -> list[Any]:
    parts = re.split(r"(\d+)", str(value or ""))
    out: list[Any] = []
    for part in parts:
        out.append(int(part) if part.isdigit() else part.lower())
    return out


def truthy(value: Any) -> bool:
    return text(value).strip().lower() in {"1", "true", "t", "yes", "y", "on"}


def bar_area_mm2(bar: int | None) -> float | None:
    if not bar or int(bar) <= 0:
        return None
    return math.pi * float(bar) * float(bar) / 4.0


def wall_schedule_area_mm2pm(item: dict[str, Any] | None) -> float:
    if not item:
        return 0.0
    bar = int(as_float((item or {}).get("bar"), 0) or 0)
    spacing = float(as_float((item or {}).get("spacing"), 0) or 0.0)
    layers = float(as_float((item or {}).get("layers"), 0) or 0.0)
    area = bar_area_mm2(bar)
    if not area or spacing <= 0.0 or layers <= 0.0:
        return 0.0
    return float(area) * 1000.0 / spacing * layers


def wall_schedule_dowel_spacing_mm(item: dict[str, Any] | None) -> int:
    if not item:
        return 0
    area = wall_schedule_area_mm2pm(item)
    dowel_size = int(as_float((item or {}).get("dowel_size"), 0) or 0)
    dowel_bar_area = bar_area_mm2(dowel_size)
    if area <= 0.0 or area > 5000.0 or not dowel_bar_area:
        return 0
    raw_spacing = 1000.0 / (area / float(dowel_bar_area))
    return max(int(math.floor(raw_spacing / 25.0) * 25.0), 0)


def wall_schedule_dowel_area_mm2pm(item: dict[str, Any] | None) -> float:
    if not item:
        return 0.0
    dowel_size = int(as_float((item or {}).get("dowel_size"), 0) or 0)
    dowel_spacing = wall_schedule_dowel_spacing_mm(item)
    dowel_bar_area = bar_area_mm2(dowel_size)
    if dowel_spacing <= 0 or not dowel_bar_area:
        return 0.0
    return float(dowel_bar_area) * 1000.0 / float(dowel_spacing)


def normalise_wall_schedule_item(item: dict[str, Any] | None) -> dict[str, Any]:
    row = deepcopy(item or {})
    row["area"] = wall_schedule_area_mm2pm(row)
    row["dowel_spacing"] = wall_schedule_dowel_spacing_mm(row)
    row["dowel_area"] = wall_schedule_dowel_area_mm2pm(row)
    return row


def spandrel_long_area_mm2(item: dict[str, Any] | None) -> float:
    if not item:
        return 0.0
    bar = int(as_float((item or {}).get("bar"), 0) or 0)
    bars = float(as_float((item or {}).get("bars"), 0) or 0.0)
    area = bar_area_mm2(bar)
    if not area or bars <= 0.0:
        return 0.0
    return float(area) * bars


def normalise_spandrel_long_schedule_item(item: dict[str, Any] | None) -> dict[str, Any]:
    row = deepcopy(item or {})
    row["area"] = spandrel_long_area_mm2(row)
    return row


def spandrel_lig_area_mm2pm(item: dict[str, Any] | None) -> float:
    if not item:
        return 0.0
    bar = int(as_float((item or {}).get("bar"), 0) or 0)
    legs = float(as_float((item or {}).get("legs"), 0) or 0.0)
    spacing = float(as_float((item or {}).get("spacing"), 0) or 0.0)
    area = bar_area_mm2(bar)
    if not area or legs <= 0.0 or spacing <= 0.0:
        return 0.0
    return float(area) * legs * 1000.0 / spacing


def normalise_spandrel_lig_schedule_item(item: dict[str, Any] | None) -> dict[str, Any]:
    row = deepcopy(item or {})
    row["area"] = spandrel_lig_area_mm2pm(row)
    return row


def bar_mass_per_m_kg(bar: int | None) -> float:
    area = bar_area_mm2(bar)
    return float(area or 0.0) * 1e-6 * STEEL_DENSITY_FACTOR if area else 0.0


def steel_mass_for_length(bar: int | None, length_m: float | None, count: float | int | None = 1.0) -> float:
    return float(count or 0.0) * float(length_m or 0.0) * bar_mass_per_m_kg(bar)


def ratio_from_value(value: Any, fc_mpa: float | None) -> float | None:
    raw = as_float(value)
    if raw is None:
        return None
    raw = float(raw)
    if abs(raw) <= 1.0:
        return raw
    if fc_mpa and float(fc_mpa) > 0.0:
        return raw / float(fc_mpa)
    return None


def first_not_none(*values: Any) -> Any:
    for value in values:
        if value is not None:
            return value
    return None


def find_sheet(workbook, target: str, required: bool = True):
    target_n = norm_name(target)
    for ws in workbook.worksheets:
        title_n = norm_name(ws.title)
        if title_n == target_n or target_n in title_n:
            return ws
    if required:
        raise KeyError(f"Sheet containing '{target}' was not found.")
    return None


def read_table(ws) -> tuple[list[str], list[str], list[dict[str, Any]]]:
    headers = [text(c.value) for c in ws[2]]
    units = [text(c.value) for c in ws[3]]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not any(v not in (None, "") for v in row):
            continue
        rows.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return headers, units, rows


def build_story_elevations(
    stories_top_down: list[str],
    height_by_story: dict[str, float | None],
    base_story_name: str | None,
    base_story_elev_m: float | None,
) -> dict[str, float]:
    elev: dict[str, float] = {}
    if not stories_top_down:
        return elev
    bottom_up = list(reversed(stories_top_down))
    if base_story_name and base_story_name in bottom_up and base_story_elev_m is not None:
        elev[base_story_name] = float(base_story_elev_m)
    else:
        elev[bottom_up[0]] = float(base_story_elev_m or 0.0)
    for i in range(1, len(bottom_up)):
        elev[bottom_up[i]] = elev[bottom_up[i - 1]] + float(height_by_story.get(bottom_up[i]) or 0.0)
    return elev


def read_base_story_definition(workbook) -> tuple[str | None, float | None]:
    ws = find_sheet(workbook, "Tower and Base Story Definition", required=False)
    if ws is None:
        return None, None
    headers, _, rows = read_table(ws)
    if not rows:
        return None, None
    bs_name_key = next((h for h in headers if norm_name(h) in {"bsname", "basestory", "basestorename"}), None)
    bs_elev_key = next((h for h in headers if norm_name(h) in {"bselev", "baseelev", "baseelevation"}), None)
    row = rows[0]
    return text(row.get(bs_name_key)), as_float(row.get(bs_elev_key))


def rotated_rect_corners(cx: float, cy: float, width: float, thickness: float, angle_deg: float) -> list[list[float]]:
    hw, ht = width / 2.0, thickness / 2.0
    pts = [(-hw, -ht), (hw, -ht), (hw, ht), (-hw, ht)]
    a = math.radians(angle_deg)
    c, s = math.cos(a), math.sin(a)
    return [[cx + x * c - y * s, cy + x * s + y * c] for x, y in pts]


def axis_from_grid_type(value: str) -> str | None:
    v = text(value).lower()
    if "x" in v:
        return "X"
    if "y" in v:
        return "Y"
    return None


def element_intersects_grid(axis: str, ordinate_m: float, corners: list[list[float]]) -> bool:
    if not corners:
        return False
    idx = 0 if axis == "X" else 1
    values = [p[idx] for p in corners]
    return min(values) - GRID_INTERSECTION_TOL_M <= ordinate_m <= max(values) + GRID_INTERSECTION_TOL_M


def alignment_to_grid(axis: str, angle_deg: float) -> float:
    a = abs(((angle_deg + 180.0) % 180.0))
    if axis == "X":
        return max(abs(math.sin(math.radians(a))), abs(math.sin(math.radians(a - 180.0))))
    return max(abs(math.cos(math.radians(a))), abs(math.cos(math.radians(a - 180.0))))


def projected_station_and_width(axis: str, item: dict[str, Any]) -> tuple[float, float]:
    coords = [p[1] for p in item["corners"]] if axis == "X" else [p[0] for p in item["corners"]]
    return (sum(coords) / len(coords), max(max(coords) - min(coords), 0.05))


def parse_reinf_string(value: Any) -> tuple[int | None, int | None]:
    m = re.match(r"^\s*N\s*(\d+)\s*-\s*(\d+)\s*$", text(value), re.I)
    if not m:
        return None, None
    return int(m.group(1)), int(m.group(2))


def area_from_reinf(value: Any) -> float | None:
    bar, spacing = parse_reinf_string(value)
    if not bar or not spacing or spacing <= 0:
        return None
    return 2.0 * (math.pi * bar * bar / 4.0) * (1000.0 / spacing)


def parse_fc_mpa(material: Any) -> float | None:
    m = re.search(r"(\d+(?:\.\d+)?)", text(material))
    return float(m.group(1)) if m else None


def fc_for_material(dataset: dict[str, Any] | None, material: Any) -> float | None:
    mat = text(material)
    if dataset:
        fc = (dataset.get("fc_by_material") or {}).get(mat)
        if fc is not None:
            return float(fc)
    return parse_fc_mpa(mat)


def hook_length_mm(bar: int | None) -> float:
    return float(HOOK_TABLE_MM.get(int(bar or 0), 0.0))


def wall_lap_length_mm(bar: int | None) -> float:
    return float(DEFAULT_WALL_LAP_TABLE_MM.get(int(bar or 0), 0.0))


def wall_dowel_lap_length_mm(bar: int | None) -> float:
    return float(DEFAULT_WALL_DOWEL_LAP_TABLE_MM.get(int(bar or 0), 0.0))


def spandrel_lap_length_mm(bar: int | None) -> float:
    return float((DEFAULT_SPANDREL_LAP_TABLE_MM.get(int(bar or 0)) or {}).get("lap") or 0.0)


def spandrel_cog_length_mm(bar: int | None) -> float:
    return float((DEFAULT_SPANDREL_LAP_TABLE_MM.get(int(bar or 0)) or {}).get("cog") or 0.0)


def bars_per_layer(span_m: float | None, spacing_mm: int | float | None) -> int | None:
    span_mm = max(float(span_m or 0.0) * 1000.0, 0.0)
    if span_mm <= 0.0 or not spacing_mm or float(spacing_mm) <= 0.0:
        return None
    return int(math.ceil(span_mm / float(spacing_mm)) + 1)


def failure_message(value: Any) -> bool:
    v = text(value).strip().lower()
    return v not in {"", "none", "no message", "ok", "0"}


def parse_material_strengths(workbook) -> dict[str, float]:
    ws = (
        find_sheet(workbook, "Mat Prop - Concrete Data", required=False)
        or find_sheet(workbook, "Materials", required=False)
        or find_sheet(workbook, "Material", required=False)
    )
    if ws is None:
        return {}
    headers, _, rows = read_table(ws)
    mat_h = next((h for h in headers if norm_name(h) == "material"), None)
    fc_h = next((h for h in headers if norm_name(h) in {"fc", "fcp", "fpc", "fprimec"}), None)
    out: dict[str, float] = {}
    for row in rows:
        mat = text(row.get(mat_h))
        fc = as_float(row.get(fc_h))
        if mat and fc is not None:
            out[mat] = float(fc)
    return out


def wall_schedule_items_for_dataset(dataset: dict[str, Any] | None = None, active_only: bool = True) -> list[dict[str, Any]]:
    items = [normalise_wall_schedule_item(item) for item in DEFAULT_WALL_REINF_SCHEDULE]
    if not active_only or dataset is None:
        return items
    active = {text(mark) for mark in dataset.get("wall_schedule_active_marks", []) if text(mark)}
    return [item for item in items if text(item.get("mark")) in active]


def spandrel_long_items_for_dataset(dataset: dict[str, Any] | None = None, active_only: bool = True) -> list[dict[str, Any]]:
    items = [normalise_spandrel_long_schedule_item(item) for item in DEFAULT_SPANDREL_LONG_SCHEDULE]
    if not active_only or dataset is None:
        return items
    active = {text(mark) for mark in dataset.get("spandrel_tb_active_marks", []) if text(mark)}
    return [item for item in items if text(item.get("mark")) in active]


def spandrel_lig_items_for_dataset(dataset: dict[str, Any] | None = None, active_only: bool = True) -> list[dict[str, Any]]:
    items = [normalise_spandrel_lig_schedule_item(item) for item in DEFAULT_SPANDREL_LIG_SCHEDULE]
    if not active_only or dataset is None:
        return items
    active = {text(mark) for mark in dataset.get("spandrel_lig_active_marks", []) if text(mark)}
    return [item for item in items if text(item.get("mark")) in active]


def choose_schedule(required_area: float | None, schedule_items: list[dict[str, Any]]) -> dict[str, Any] | None:
    if required_area is None or required_area <= 0:
        return None
    items = [deepcopy(item) for item in schedule_items if float(item.get("area") or 0.0) > 0.0]
    if not items:
        return None
    for item in items:
        if float(item.get("area") or 0.0) >= float(required_area):
            return deepcopy(item)
    return deepcopy(items[-1])


def wall_schedule_text(item: dict[str, Any] | None) -> str:
    if not item:
        return "—"
    return f"N{int(item['bar'])}-{int(item['spacing'])}"


def wall_schedule_lookup_by_mark(mark: Any, dataset: dict[str, Any] | None = None) -> dict[str, Any] | None:
    target = text(mark)
    for item in wall_schedule_items_for_dataset(dataset, active_only=False):
        if text(item.get("mark")) == target:
            return deepcopy(item)
    return None


def wall_schedule_lookup_by_reinf(value: Any, dataset: dict[str, Any] | None = None) -> dict[str, Any] | None:
    bar, spacing = parse_reinf_string(value)
    if not bar or not spacing:
        return None
    for item in wall_schedule_items_for_dataset(dataset, active_only=False):
        if int(item.get("bar") or 0) == int(bar) and int(item.get("spacing") or 0) == int(spacing):
            return deepcopy(item)
    return None


def spandrel_long_lookup_by_mark(mark: Any, dataset: dict[str, Any] | None = None) -> dict[str, Any] | None:
    target = text(mark)
    for item in spandrel_long_items_for_dataset(dataset, active_only=False):
        if text(item.get("mark")) == target:
            return deepcopy(item)
    return None


def spandrel_lig_lookup_by_mark(mark: Any, dataset: dict[str, Any] | None = None) -> dict[str, Any] | None:
    target = text(mark)
    for item in spandrel_lig_items_for_dataset(dataset, active_only=False):
        if text(item.get("mark")) == target:
            return deepcopy(item)
    return None


def spandrel_long_display_text(item: dict[str, Any] | None) -> str:
    if not item:
        return "—"
    return f"{int(item['bars'])}-N{int(item['bar'])}"


def spandrel_lig_display_text(item: dict[str, Any] | None) -> str:
    if not item:
        return "—"
    return f"{int(item['legs'])}L-N{int(item['bar'])}-{int(item['spacing'])}"


def current_reinf_value(override_value: str | None, fallback: str | None) -> str:
    return text(override_value) or text(fallback) or "—"


def reinforcement_strings(base: dict[str, Any], override: dict[str, Any] | None) -> tuple[str, str, str]:
    vertical = current_reinf_value((override or {}).get("vertical"), base.get("seed_vertical"))
    horizontal = current_reinf_value((override or {}).get("horizontal"), base.get("seed_horizontal"))
    return vertical, horizontal, f"V: {vertical} / H: {horizontal}"


def make_item(kind: str, label: str, source_story: str, plot_story: str, material: str, angle_deg: float, cgx_m: float, cgy_m: float, width_m: float, thickness_m: float, depth_m: float = 0.0) -> dict[str, Any]:
    return {
        "key": f"{kind}|{source_story}|{label}",
        "kind": kind,
        "label": label,
        "source_story": source_story,
        "plot_story": plot_story,
        "material": material,
        "angle_deg": float(angle_deg or 0.0),
        "cgx_m": float(cgx_m or 0.0),
        "cgy_m": float(cgy_m or 0.0),
        "width_m": float(width_m or 0.0),
        "thickness_m": float(thickness_m or 0.0),
        "depth_m": float(depth_m or 0.0),
        "corners": rotated_rect_corners(float(cgx_m or 0.0), float(cgy_m or 0.0), float(width_m or 0.0), float(thickness_m or 0.0), float(angle_deg or 0.0)),
    }


def parse_pier_design_sheet(ws, walls: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    _, _, rows = read_table(ws)
    wall_lookup = {(w["source_story"], w["label"]): w for w in walls}
    grouped: dict[str, dict[str, Any]] = {}
    for row in rows:
        story = text(row.get("Story"))
        label = text(row.get("Pier Label") or row.get("Pier"))
        if not story or not label:
            continue
        wall = wall_lookup.get((story, label))
        thickness_mm = as_float(row.get("Thickness"))
        if thickness_mm is None and wall:
            thickness_mm = float((wall or {}).get("thickness_m") or 0.0) * 1000.0
        req_pct = as_float(row.get("Required Reinf. Percentage"))
        cur_pct = as_float(row.get("Current Reinf. Percentage") or row.get("Actual %"))
        shear = as_float(row.get("Shear Rebar"))
        key = f"wall|{story}|{label}"
        entry = grouped.setdefault(key, {"messages": []})
        if req_pct is not None:
            entry["required_vertical_pct"] = max(float(entry.get("required_vertical_pct") or 0.0), float(req_pct))
        if cur_pct is not None:
            entry["current_vertical_pct"] = max(float(entry.get("current_vertical_pct") or 0.0), float(cur_pct))
        if thickness_mm is not None and req_pct is not None:
            required_v = (float(req_pct) / 100.0) * float(thickness_mm) * 1000.0
            entry["required_vertical_mm2pm"] = max(float(entry.get("required_vertical_mm2pm") or 0.0), required_v)
        if thickness_mm is not None and cur_pct is not None:
            current_v = (float(cur_pct) / 100.0) * float(thickness_mm) * 1000.0
            entry["current_vertical_mm2pm"] = max(float(entry.get("current_vertical_mm2pm") or 0.0), current_v)
        if shear is not None:
            entry["required_horizontal_mm2pm"] = max(float(entry.get("required_horizontal_mm2pm") or 0.0), float(shear))
        if failure_message(row.get("Warnings")) or failure_message(row.get("Errors")):
            entry["has_failure"] = True
            msg = " / ".join([m for m in [text(row.get("Warnings")), text(row.get("Errors"))] if failure_message(m)])
            if msg:
                entry.setdefault("messages", []).append(msg)
    return grouped


def find_first_header(headers: list[str], patterns: list[str]) -> str | None:
    normalized = {header: norm_name(header) for header in headers if text(header)}
    for pattern in patterns:
        pattern_n = norm_name(pattern)
        for header, header_n in normalized.items():
            if header_n == pattern_n or pattern_n in header_n:
                return header
    return None


def parse_boundary_workbook(path: Path) -> dict[str, dict[str, Any]]:
    wb = load_workbook(path, data_only=True, read_only=False)
    ws = find_sheet(wb, "Pier Dgn Sum")
    headers, _, rows = read_table(ws)
    story_header = find_first_header(headers, ["Story", "Pier Story"])
    pier_header = find_first_header(headers, ["Pier Label", "Pier", "Pier Name"])
    left_header = find_first_header(headers, [
        "Boundary Zone Left", "Boundary Left", "Boundary Length Left", "Left Boundary Length",
        "Boundary Zone Length Left", "Length of Left Boundary", "Left Boundary Zone", "Left Boundary",
        "Boundary Element Left", "Left Boundary Element", "BZ Left", "Left BZ", "BE Left", "Left BE",
    ])
    right_header = find_first_header(headers, [
        "Boundary Zone Right", "Boundary Right", "Boundary Length Right", "Right Boundary Length",
        "Boundary Zone Length Right", "Length of Right Boundary", "Right Boundary Zone", "Right Boundary",
        "Boundary Element Right", "Right Boundary Element", "BZ Right", "Right BZ", "BE Right", "Right BE",
    ])
    left_stress_header = find_first_header(headers, [
        "Left Compressive Stress", "Compressive Stress Left", "Left Compression Stress", "Compression Stress Left",
        "Left Comp Stress", "Comp Stress Left", "Boundary Compressive Stress Left", "Left Boundary Stress",
        "BE Stress Left", "Stress Left"
    ])
    right_stress_header = find_first_header(headers, [
        "Right Compressive Stress", "Compressive Stress Right", "Right Compression Stress", "Compression Stress Right",
        "Right Comp Stress", "Comp Stress Right", "Boundary Compressive Stress Right", "Right Boundary Stress",
        "BE Stress Right", "Stress Right"
    ])
    left_limit_header = find_first_header(headers, [
        "Left Compression Limit", "Compression Limit Left", "Left Compressive Limit", "Compressive Limit Left",
        "Left Stress Limit", "Stress Limit Left", "Boundary Compression Limit Left", "BE Limit Left", "Limit Left"
    ])
    right_limit_header = find_first_header(headers, [
        "Right Compression Limit", "Compression Limit Right", "Right Compressive Limit", "Compressive Limit Right",
        "Right Stress Limit", "Stress Limit Right", "Boundary Compression Limit Right", "BE Limit Right", "Limit Right"
    ])
    if not story_header or not pier_header or (not left_header and not right_header):
        raise KeyError("Could not identify boundary-zone columns in 'Pier Dgn Sum'.")
    out: dict[str, dict[str, Any]] = {}
    for row in rows:
        story = text(row.get(story_header))
        label = text(row.get(pier_header))
        if not story or not label:
            continue
        key = f"wall|{story}|{label}"
        entry = out.setdefault(key, {})
        left_mm = max(float(as_float(row.get(left_header), 0.0) or 0.0), 0.0) if left_header else 0.0
        right_mm = max(float(as_float(row.get(right_header), 0.0) or 0.0), 0.0) if right_header else 0.0
        entry["left_mm"] = max(float(entry.get("left_mm") or 0.0), left_mm)
        entry["right_mm"] = max(float(entry.get("right_mm") or 0.0), right_mm)
        ls = as_float(row.get(left_stress_header)) if left_stress_header else None
        rs = as_float(row.get(right_stress_header)) if right_stress_header else None
        ll = as_float(row.get(left_limit_header)) if left_limit_header else None
        rl = as_float(row.get(right_limit_header)) if right_limit_header else None
        if ls is not None:
            entry["left_stress_raw"] = max(float(entry.get("left_stress_raw") or 0.0), float(ls))
        if rs is not None:
            entry["right_stress_raw"] = max(float(entry.get("right_stress_raw") or 0.0), float(rs))
        if ll is not None:
            entry["left_limit_raw"] = max(float(entry.get("left_limit_raw") or 0.0), float(ll))
        if rl is not None:
            entry["right_limit_raw"] = max(float(entry.get("right_limit_raw") or 0.0), float(rl))
        if (
            entry.get("left_mm", 0.0) <= 0.0
            and entry.get("right_mm", 0.0) <= 0.0
            and all(float(entry.get(k) or 0.0) <= 0.0 for k in ("left_stress_raw", "right_stress_raw", "left_limit_raw", "right_limit_raw"))
        ):
            out.pop(key, None)
    return out


def boundary_hatch_style(fc_mpa: float | None, stress_ratio: float | None) -> str | None:
    fc = float(fc_mpa or 0.0)
    if stress_ratio is None:
        return "diag-red"
    if stress_ratio > 0.2:
        return "diag-blue"
    if 0.15 < stress_ratio < 0.2:
        return "dot-red" if fc <= 50.0 else "dot-blue"
    return "diag-red"


def body_hatch_style(fc_mpa: float | None) -> str | None:
    fc = float(fc_mpa or 0.0)
    if fc > 65.0:
        return "diag-blue"
    if fc > 50.0:
        return "dot-blue"
    return None


def wall_horizontal_splice_factor(length_m: float | None) -> float:
    L = float(length_m or 0.0)
    for limit, factor in DEFAULT_WALL_HORIZONTAL_SPLICE_FACTORS:
        if L <= float(limit):
            return float(factor)
    return 5.0


def wall_minimum_pct(fc_mpa: float | None, zone: str, red_level: int = 0) -> float:
    if zone == "NO":
        return 0.0
    base_pct = max(0.25, 100.0 * 0.7 * math.sqrt(max(float(fc_mpa or 0.0), 0.0)) / 500.0)
    if zone == "FULL":
        return base_pct
    reduced_pct = base_pct * max(0.0, 1.0 - 0.1 * max(int(red_level or 0), 0))
    return max(0.25, reduced_pct)


def build_ligature_data(item: dict[str, Any], dataset: dict[str, Any]) -> dict[str, Any]:
    boundary = deepcopy(dataset.get("boundary_by_key", {}).get(item.get("key"), {}))
    left_mm = max(float(boundary.get("left_mm") or 0.0), 0.0)
    right_mm = max(float(boundary.get("right_mm") or 0.0), 0.0)
    fc_mpa = float(item.get("fc_mpa") or fc_for_material(dataset, item.get("material")) or 0.0)
    left_stress_ratio = ratio_from_value(boundary.get("left_stress_raw"), fc_mpa)
    right_stress_ratio = ratio_from_value(boundary.get("right_stress_raw"), fc_mpa)
    left_limit_ratio = ratio_from_value(boundary.get("left_limit_raw"), fc_mpa)
    right_limit_ratio = ratio_from_value(boundary.get("right_limit_raw"), fc_mpa)
    has_boundary = left_mm > 0.0 or right_mm > 0.0

    vertical_item = item.get("current_vertical_item") or wall_schedule_lookup_by_reinf(item.get("current_vertical"), dataset)
    horizontal_item = item.get("current_horizontal_item") or wall_schedule_lookup_by_reinf(item.get("current_horizontal"), dataset)
    vertical_bar = int((vertical_item or {}).get("bar") or 0)
    horizontal_spacing = int((horizontal_item or {}).get("spacing") or 0)

    ast_v_pct = None
    if vertical_item and item.get("thickness_m"):
        ast_v_pct = 100.0 * float(vertical_item.get("area") or 0.0) / (float(item.get("thickness_m") or 0.0) * 1_000_000.0)
    ast_h_pct = None
    if horizontal_item and item.get("thickness_m"):
        ast_h_pct = 100.0 * float(horizontal_item.get("area") or 0.0) / (float(item.get("thickness_m") or 0.0) * 1_000_000.0)

    required_vertical_pct = as_float(item.get("required_vertical_pct"))
    type1_required = not (fc_mpa <= 50.0 and (required_vertical_pct is not None and float(required_vertical_pct) < 1.0) and (ast_h_pct is not None and float(ast_h_pct) >= 0.25))
    type2_required = fc_mpa > 50.0

    L_mm = float(item.get("width_m") or 0.0) * 1000.0
    if ast_v_pct is None:
        duct3_bound_mm = 0.0
    elif ast_v_pct <= 0.75:
        duct3_bound_mm = 0.15 * L_mm
    elif ast_v_pct <= 1.5:
        duct3_bound_mm = 0.18 * L_mm
    elif ast_v_pct <= 2.5:
        duct3_bound_mm = 0.20 * L_mm
    elif ast_v_pct <= 3.2:
        duct3_bound_mm = 0.25 * L_mm
    else:
        duct3_bound_mm = 0.50 * L_mm

    ductility_level = str((dataset.get("settings") or {}).get("ductility_level") or DEFAULT_DUCTILITY_LEVEL)
    boundary_sum_mm = left_mm + right_mm
    if ductility_level == "1":
        boundary_reg_mm = 0.0
    elif ductility_level == "2":
        boundary_reg_mm = boundary_sum_mm
    else:
        boundary_reg_mm = 2.0 * duct3_bound_mm if boundary_sum_mm > 0.0 else 0.0
    boundary_reg_mm = 50.0 * math.ceil(boundary_reg_mm / 50.0) if boundary_reg_mm > 0.0 else 0.0

    lig_bar = int((vertical_item or {}).get("lig_bar") or 0)
    lig_spacing1 = int((horizontal_item or {}).get("spacing") or 0)
    default_spacing2 = int((vertical_item or {}).get("lig_spacing1") or 0)
    boundary_spacing2 = int((vertical_item or {}).get("lig_spacing2") or 0)
    raw_spacing2 = float(boundary_spacing2 if (type2_required or boundary_reg_mm > 0.0) else default_spacing2)
    tw_mm = float(item.get("thickness_m") or 0.0) * 1000.0
    cap1 = tw_mm / (1.0 if not type2_required else 2.0) if tw_mm > 0.0 else 0.0
    cap2 = tw_mm * 0.8 if fc_mpa > 50.0 and tw_mm > 0.0 else 99999.0
    lig_spacing2 = min(v for v in [raw_spacing2, cap1, cap2] if v and v > 0.0) if raw_spacing2 > 0.0 else 0.0

    H = float(dataset.get("height_by_story", {}).get(item.get("source_story")) or 0.0)
    extent_m = float(item.get("width_m") or 0.0) if type1_required else boundary_reg_mm / 1000.0
    no_ligs = math.ceil((extent_m * H * 1_000_000.0) / float(lig_spacing1) / float(lig_spacing2)) if extent_m > 0.0 and H > 0.0 and lig_spacing1 > 0.0 and lig_spacing2 > 0.0 else 0
    cover_mm = float((dataset.get("settings") or {}).get("cover_mm") or DEFAULT_COVER_MM)
    lig_length_m = max((tw_mm - 2.0 * cover_mm) / 1000.0 + 2.0 * 170.0 / 1000.0, 0.0) if tw_mm > 0.0 else 0.0

    base = {
        "has_boundary": has_boundary,
        "boundary_left_mm": left_mm,
        "boundary_right_mm": right_mm,
        "boundary_left_stress_ratio": left_stress_ratio,
        "boundary_right_stress_ratio": right_stress_ratio,
        "boundary_left_limit_ratio": left_limit_ratio,
        "boundary_right_limit_ratio": right_limit_ratio,
        "boundary_reg_mm": boundary_reg_mm,
        "duct3_bound_mm": duct3_bound_mm,
        "type1_ligs_required": type1_required,
        "type2_ligs_required": type2_required,
        "ligs_bar_mm": lig_bar or None,
        "ligs_spacing1_mm": lig_spacing1 or None,
        "ligs_spacing2_mm": lig_spacing2 or None,
        "extent_m": extent_m if extent_m > 0.0 else None,
        "horizontal_spacing_mm": horizontal_item.get("spacing") if horizontal_item else None,
        "horizontal_spacing_warning": bool(has_boundary and horizontal_item and float(horizontal_item.get("spacing") or 0.0) > 200.0),
    }
    if no_ligs <= 0 or lig_bar <= 0 or lig_length_m <= 0.0:
        return {**base, "ligs": "—", "ligs_n": None, "ligs_sets": None, "ligs_length_m": None}
    ligs_text = f"{int(no_ligs)}xN{lig_bar}"
    return {**base, "ligs": ligs_text, "ligs_n": int(no_ligs), "ligs_sets": None, "ligs_length_m": lig_length_m}


def wall_quantities(item: dict[str, Any], dataset: dict[str, Any]) -> dict[str, float | int | None]:
    if item.get("kind") != "wall":
        return {
            "steel_kg": None, "concrete_volume_m3": None, "rate_kgpm3": None,
            "vertical_kg": None, "horizontal_kg": None, "ligs_kg": None,
            "dowels_kg": None, "ubars_kg": None,
            "vertical_bar_count_per_layer": None, "horizontal_bar_count_per_layer": None,
        }
    L = max(float(item.get("width_m") or 0.0), 0.0)
    t = max(float(item.get("thickness_m") or 0.0), 0.0)
    H = max(float(dataset.get("height_by_story", {}).get(item.get("source_story")) or 0.0), 0.0)
    volume = L * t * H if L > 0 and t > 0 and H > 0 else None
    fc = float(item.get("fc_mpa") or fc_for_material(dataset, item.get("material")) or 0.0)

    vertical_item = item.get("current_vertical_item") or wall_schedule_lookup_by_reinf(item.get("current_vertical"), dataset)
    horizontal_item = item.get("current_horizontal_item") or wall_schedule_lookup_by_reinf(item.get("current_horizontal"), dataset)
    v_bar = int((vertical_item or {}).get("bar") or 0)
    h_bar = int((horizontal_item or {}).get("bar") or 0)
    v_spacing = int((vertical_item or {}).get("spacing") or 0)
    h_spacing = int((horizontal_item or {}).get("spacing") or 0)

    vertical_bar_count_per_layer = bars_per_layer(L, v_spacing)
    horizontal_bar_count_per_layer = bars_per_layer(H, h_spacing)

    ast_v_eff = float((vertical_item or {}).get("area") or 0.0)
    ast_h_eff = float((horizontal_item or {}).get("area") or 0.0)

    wall_type = text(item.get("wall_type") or DEFAULT_WALL_TYPE) or DEFAULT_WALL_TYPE
    zone = text(item.get("ductility_zone") or "NO")
    lap_factor_v = 1.0 if wall_type == "PRECAST" else (1.5 if str((dataset.get("settings") or {}).get("ductility_level")) == "3" and zone == "FULL" else 1.0)
    lap_ratio_v = (lap_factor_v * (wall_lap_length_mm(v_bar) / 1000.0 + H) / H) if v_bar and H > 0 else None
    lap_ratio_h = ((wall_lap_length_mm(h_bar) / 1000.0 * wall_horizontal_splice_factor(L)) + L) / L if h_bar and L > 0 else None

    vertical_kg = ast_v_eff * L * H * 7850.0 / 1_000_000.0 * float(lap_ratio_v or 0.0) if ast_v_eff > 0 and L > 0 and H > 0 and lap_ratio_v else None
    horizontal_kg = ast_h_eff * L * H * 7850.0 / 1_000_000.0 * float(lap_ratio_h or 0.0) if ast_h_eff > 0 and L > 0 and H > 0 and lap_ratio_h else None

    ligs_bar = int(item.get("ligs_bar_mm") or 0)
    ligs_n = int(item.get("ligs_n") or 0)
    ligs_length_m = float(item.get("ligs_length_m") or 0.0)
    ligs_kg = ligs_n * ligs_length_m * (bar_area_mm2(ligs_bar) or 0.0) * 7850.0 / 1_000_000.0 if ligs_n > 0 and ligs_bar > 0 and ligs_length_m > 0 else None

    global_wall_type = text((dataset.get("settings") or {}).get("wall_type") or DEFAULT_WALL_TYPE)
    dowel_size = int((vertical_item or {}).get("dowel_size") or 0) if global_wall_type == "PRECAST" else 0
    dowel_spacing = int((vertical_item or {}).get("dowel_spacing") or 0) if global_wall_type == "PRECAST" else 0
    no_dowels = int(math.ceil(L * 1000.0 / dowel_spacing + 1.0)) if dowel_spacing > 0 and L > 0 else 0
    dowel_len_m = (2.0 * wall_dowel_lap_length_mm(dowel_size) / 1000.0) * (1.5 if str((dataset.get("settings") or {}).get("ductility_level")) == "3" else 1.0) if dowel_size > 0 and dowel_spacing > 0 else 0.0
    dowels_kg = no_dowels * dowel_len_m * (bar_area_mm2(dowel_size) or 0.0) * 7850.0 / 1_000_000.0 if no_dowels > 0 and dowel_len_m > 0.0 else None

    ubar_size = 10 if t <= 0.25 else 12
    no_ubars = 2 * int(math.ceil(1.0 + L / (v_spacing / 1000.0))) if dowel_size > 0 and v_spacing > 0 and L > 0 else 0
    cover_mm = float((dataset.get("settings") or {}).get("cover_mm") or DEFAULT_COVER_MM)
    ubar_len_m = max(t - 2.0 * cover_mm / 1000.0 + 2.0 * min(0.8, dowel_len_m / 2.0), 0.0) if dowel_size > 0 else 0.0
    ubars_kg = no_ubars * ubar_len_m * (bar_area_mm2(ubar_size) or 0.0) * 7850.0 / 1_000_000.0 if no_ubars > 0 and ubar_len_m > 0.0 else None

    steel_components = [value for value in (vertical_kg, horizontal_kg, ligs_kg, dowels_kg, ubars_kg) if value is not None]
    steel_kg = sum(steel_components) if steel_components else None
    rate = (steel_kg / volume) if (steel_kg is not None and volume and volume > 0) else None
    return {
        "steel_kg": steel_kg,
        "concrete_volume_m3": volume,
        "rate_kgpm3": rate,
        "vertical_kg": vertical_kg,
        "horizontal_kg": horizontal_kg,
        "ligs_kg": ligs_kg,
        "dowels_kg": dowels_kg,
        "ubars_kg": ubars_kg,
        "vertical_bar_count_per_layer": vertical_bar_count_per_layer,
        "horizontal_bar_count_per_layer": horizontal_bar_count_per_layer,
        "dowel_size": dowel_size or None,
        "dowel_spacing": dowel_spacing or None,
        "no_dowels": no_dowels or None,
        "dowel_length_m": dowel_len_m or None,
        "ubar_size": ubar_size if dowel_size > 0 else None,
        "no_ubars": no_ubars or None,
        "ubar_length_m": ubar_len_m or None,
    }


def parse_spandrel_design_sheet(ws, spandrels: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    _, _, rows = read_table(ws)
    lookup = {(s["source_story"], s["label"]): s for s in spandrels}
    grouped: dict[str, dict[str, Any]] = {}
    for row in rows:
        story = text(row.get("Story"))
        label = text(row.get("Spandrel"))
        if not story or not label:
            continue
        base = lookup.get((story, label))
        if not base:
            continue
        key = f"spandrel|{story}|{label}"
        entry = grouped.setdefault(key, {"messages": []})
        top_rebar = max(float(as_float(row.get("Top Rebar"), 0.0) or 0.0), 0.0)
        bottom_rebar = max(float(as_float(row.get("Bottom Rebar"), 0.0) or 0.0), 0.0)
        av_vert = max(float(as_float(row.get("Av Vert"), 0.0) or 0.0), 0.0)
        design_shear = float(as_float(row.get("Design Shear"), 0.0) or 0.0)
        top_m = float(as_float(row.get("Top Design Moment"), 0.0) or 0.0)
        bottom_m = float(as_float(row.get("Bottom Design Moment"), 0.0) or 0.0)
        H_mm = float(base.get("width_m") or 0.0) * 1000.0
        D_mm = float(base.get("depth_m") or 0.0) * 1000.0
        fixity = "FR" if text(base.get("fixity") or "").upper().endswith("FR") or text(label).upper().endswith("FR") else "FF"
        moment_ref = max(abs(top_m), abs(bottom_m), 1e-9)
        ratio = max(1.0, ((design_shear * H_mm / 1000.0) if fixity == "FR" else (design_shear * H_mm / 2.0 / 1000.0)) / moment_ref) if H_mm > 0 else 1.0
        ast_h = max(top_rebar, bottom_rebar) * ratio
        entry["required_top_rebar_mm2"] = max(float(entry.get("required_top_rebar_mm2") or 0.0), top_rebar)
        entry["required_bottom_rebar_mm2"] = max(float(entry.get("required_bottom_rebar_mm2") or 0.0), bottom_rebar)
        entry["required_tb_area_mm2"] = max(float(entry.get("required_tb_area_mm2") or 0.0), ast_h)
        entry["required_lig_area_base_mm2pm"] = max(float(entry.get("required_lig_area_base_mm2pm") or 0.0), av_vert)
        entry["fixity"] = fixity
        if failure_message(row.get("Warnings")) or failure_message(row.get("Errors")):
            entry["has_failure"] = True
            msg = " / ".join([m for m in [text(row.get("Warnings")), text(row.get("Errors"))] if failure_message(m)])
            if msg:
                entry.setdefault("messages", []).append(msg)
    return grouped


def parse_spandrels(rows: list[dict[str, Any]], below_story_by_source: dict[str, str]) -> list[dict[str, Any]]:
    out = []
    for row in rows:
        source_story = text(row.get("Story"))
        plot_story = below_story_by_source.get(source_story)
        label = text(row.get("Spandrel"))
        if not source_story or not plot_story or not label:
            continue
        lx, ly = as_float(row.get("CG Left X")), as_float(row.get("CG Left Y"))
        rx, ry = as_float(row.get("CG Right X")), as_float(row.get("CG Right Y"))
        if None in (lx, ly, rx, ry):
            continue
        dx, dy = float(rx) - float(lx), float(ry) - float(ly)
        endpoint_len = math.hypot(dx, dy)
        length_m = endpoint_len if endpoint_len > 1e-6 else (as_float(row.get("Length"), 0.0) or 0.0) / 1000.0
        if length_m <= 0:
            continue
        angle_deg = math.degrees(math.atan2(dy, dx)) if endpoint_len > 1e-6 else 0.0
        thickness_m = max(((as_float(row.get("Thickness Left"), 0.0) or 0.0) + (as_float(row.get("Thickness Right"), 0.0) or 0.0)) / 2000.0, (as_float(row.get("Thickness Left"), 0.0) or 0.0) / 1000.0, (as_float(row.get("Thickness Right"), 0.0) or 0.0) / 1000.0)
        depth_m = max(((as_float(row.get("Depth Left"), 0.0) or 0.0) + (as_float(row.get("Depth Right"), 0.0) or 0.0)) / 2000.0, (as_float(row.get("Depth Left"), 0.0) or 0.0) / 1000.0, (as_float(row.get("Depth Right"), 0.0) or 0.0) / 1000.0)
        item = make_item("spandrel", label, source_story, plot_story, text(row.get("Material")), angle_deg, (float(lx) + float(rx)) / 2.0, (float(ly) + float(ry)) / 2.0, length_m, thickness_m, depth_m)
        item["line_x1_m"], item["line_y1_m"], item["line_x2_m"], item["line_y2_m"] = float(lx), float(ly), float(rx), float(ry)
        item["fixity"] = "FR" if text(label).upper().endswith("FR") else "FF"
        out.append(item)
    return out


def build_ductility_profile(dataset: dict[str, Any], ductility_level: str | int | None) -> dict[str, Any]:
    level = str(ductility_level or (dataset.get("settings") or {}).get("ductility_level") or "1")
    profile: dict[str, Any] = {
        "active": level in {"2", "3"},
        "ductility_level": level,
        "rows": [],
        "min_pct_by_plot_story": {},
        "fc_max_mpa": None,
        "lw_base_m": None,
        "lw_max_m": None,
        "base_pct": None,
        "threshold_height_m": None,
        "governing_height_m": None,
        "governing_condition": None,
        "second_storey_height_m": None,
        "first_source_story": None,
    }
    plot_stories = list(dataset.get("available_plot_stories", []))
    fc_values = [fc_for_material(dataset, w.get("material")) for w in dataset.get("walls", [])]
    fc_values = [float(v) for v in fc_values if v is not None]
    if not plot_stories or not fc_values:
        return profile
    first_plot_story = plot_stories[0]
    first_source_story = text(dataset.get("source_story_by_plot", {}).get(first_plot_story))
    base_walls = [w for w in dataset.get("walls", []) if text(w.get("source_story")) == first_source_story]
    lw_base = max((float(w.get("width_m") or 0.0) for w in base_walls), default=0.0)
    second_storey_height = float(dataset.get("height_by_story", {}).get(first_source_story) or 0.0)
    fc_max = max(fc_values)
    base_pct = max(0.25, 100.0 * 0.7 * math.sqrt(fc_max) / 500.0)
    threshold = max(2.0 * lw_base, second_storey_height)
    rows = []
    min_by: dict[str, float] = {}
    cumulative_h = 0.0
    red_level = 0
    for plot_story in plot_stories:
        zone = "NO"
        current_red_level = 0
        if level in {"2", "3"}:
            zone = "FULL" if cumulative_h <= threshold + 1e-9 else "RED"
            if zone == "RED":
                red_level += 1
                current_red_level = red_level
        min_pct = wall_minimum_pct(fc_max, zone, current_red_level)
        rows.append({
            "plot_story": plot_story,
            "zone": zone,
            "min_pct": min_pct,
            "cumulative_height_m": cumulative_h,
            "red_level": current_red_level,
        })
        min_by[plot_story] = min_pct
        source_story = text(dataset.get("source_story_by_plot", {}).get(plot_story))
        cumulative_h += float(dataset.get("height_by_story", {}).get(source_story) or 0.0)
    governing_height = threshold
    governing_condition = "2Lw" if 2.0 * lw_base >= second_storey_height - 1e-9 else "First storey above base"
    profile.update({
        "rows": rows,
        "min_pct_by_plot_story": min_by,
        "fc_max_mpa": fc_max,
        "lw_base_m": lw_base,
        "lw_max_m": lw_base,
        "base_pct": base_pct,
        "threshold_height_m": threshold,
        "governing_height_m": governing_height,
        "governing_condition": governing_condition,
        "second_storey_height_m": second_storey_height,
        "first_source_story": first_source_story,
    })
    return profile


def parse_workbook(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True, read_only=False)
    story_ws = find_sheet(wb, "Story Definitions")
    pier_ws = find_sheet(wb, "Pier Section Properties")
    grid_ws = find_sheet(wb, "Grid Definitions - Grid Lines", required=False)
    spandrel_ws = find_sheet(wb, "Spandrel Section Properties", required=False)
    pier_dgn_ws = find_sheet(wb, "Pier Dgn Sum", required=False)
    span_dgn_ws = find_sheet(wb, "Span Dgn Sum", required=False)
    _, _, story_rows = read_table(story_ws)
    _, _, pier_rows = read_table(pier_ws)
    _, _, grid_rows = read_table(grid_ws) if grid_ws else ([], [], [])
    _, _, spandrel_rows = read_table(spandrel_ws) if spandrel_ws else ([], [], [])
    base_story_name, base_story_elev_m = read_base_story_definition(wb)
    fc_by_material = parse_material_strengths(wb)

    stories_top_down = [text(r.get("Name")) for r in story_rows if text(r.get("Name"))]
    if base_story_name and base_story_name not in stories_top_down:
        stories_top_down = stories_top_down + [base_story_name]
    height_by_story = {text(r.get("Name")): as_float(r.get("Height")) for r in story_rows if text(r.get("Name"))}
    if base_story_name and base_story_name not in height_by_story:
        height_by_story[base_story_name] = 0.0
    story_elevation_m = build_story_elevations(stories_top_down, height_by_story, base_story_name, base_story_elev_m)
    below_story_by_source = {stories_top_down[i]: stories_top_down[i + 1] for i in range(len(stories_top_down) - 1)}
    source_story_by_plot = {v: k for k, v in below_story_by_source.items()}

    walls = []
    for row in pier_rows:
        source_story = text(row.get("Story"))
        label = text(row.get("Pier"))
        plot_story = below_story_by_source.get(source_story)
        if not source_story or not label or not plot_story:
            continue
        walls.append(make_item(
            "wall", label, source_story, plot_story, text(row.get("Material")),
            as_float(row.get("AxisAngle"), 0.0) or 0.0,
            as_float(row.get("CG Bottom X"), 0.0) or 0.0,
            as_float(row.get("CG Bottom Y"), 0.0) or 0.0,
            (as_float(row.get("Width Bottom"), 0.0) or 0.0) / 1000.0,
            (as_float(row.get("Thickness Bottom"), 0.0) or 0.0) / 1000.0,
            0.0
        ))
    spandrels = parse_spandrels(spandrel_rows, below_story_by_source) if spandrel_rows else []

    design_by_key: dict[str, dict[str, Any]] = {}
    if pier_dgn_ws:
        design_by_key.update(parse_pier_design_sheet(pier_dgn_ws, walls))
    if span_dgn_ws:
        design_by_key.update(parse_spandrel_design_sheet(span_dgn_ws, spandrels))

    pier_failure_count = sum(1 for w in walls if (design_by_key.get(w["key"], {}) or {}).get("has_failure"))
    spandrel_failure_count = sum(1 for s in spandrels if (design_by_key.get(s["key"], {}) or {}).get("has_failure"))

    systems: dict[str, list[dict[str, Any]]] = {}
    for row in grid_rows:
        system = text(row.get("Name"))
        if not system or text(row.get("Visible") or "Yes").lower() == "no":
            continue
        axis = axis_from_grid_type(text(row.get("Grid Line Type")))
        ordinate = as_float(row.get("Ordinate"))
        grid_id = text(row.get("ID"))
        if not axis or ordinate is None or not grid_id:
            continue
        systems.setdefault(system, []).append({"key": f"{system}|{axis}|{grid_id}", "system": system, "id": grid_id, "axis": axis, "ordinate_m": float(ordinate), "label": grid_id})
    core_grids = systems.get("CORE", [])
    core_grids.sort(key=lambda g: natural_sort_key(g["id"]))

    stories_bottom_up = list(reversed(stories_top_down))
    available_plot_stories = [story for story in stories_bottom_up if any(el["plot_story"] == story for el in (walls + spandrels))]
    dataset = {
        "filename": path.name,
        "stories_top_down": stories_top_down,
        "available_plot_stories": available_plot_stories,
        "grouping_story_list": stories_bottom_up,
        "height_by_story": height_by_story,
        "story_elevation_m": story_elevation_m,
        "source_story_by_plot": source_story_by_plot,
        "base_story_name": base_story_name,
        "base_story_elev_m": base_story_elev_m,
        "grids": core_grids,
        "default_grid_key": core_grids[0]["key"] if core_grids else None,
        "walls": walls,
        "spandrels": spandrels,
        "design_by_key": design_by_key,
        "overrides": {},
        "boundary_workbook": None,
        "boundary_by_key": {},
        "fc_by_material": fc_by_material,
        "settings": {"ductility_level": DEFAULT_DUCTILITY_LEVEL, "cover_mm": DEFAULT_COVER_MM, "wall_type": DEFAULT_WALL_TYPE},
        "wall_schedule_active_marks": [text(item["mark"]) for item in DEFAULT_WALL_REINF_SCHEDULE if bool(item.get("use"))],
        "spandrel_tb_active_marks": [text(item["mark"]) for item in DEFAULT_SPANDREL_LONG_SCHEDULE if bool(item.get("use"))],
        "spandrel_lig_active_marks": [text(item["mark"]) for item in DEFAULT_SPANDREL_LIG_SCHEDULE if bool(item.get("use"))],
        "pier_failure_count": pier_failure_count,
        "spandrel_failure_count": spandrel_failure_count,
        "has_pier_design": bool(pier_dgn_ws),
        "has_spandrels": bool(spandrels),
    }
    dataset["ductility_profile"] = build_ductility_profile(dataset, DEFAULT_DUCTILITY_LEVEL)
    return dataset


def wall_seed_from_design(required_area: float | None, dataset: dict[str, Any]) -> tuple[str, dict[str, Any] | None]:
    item = choose_schedule(required_area, wall_schedule_items_for_dataset(dataset))
    return wall_schedule_text(item), item


def spandrel_seed_from_design(required_area: float | None, dataset: dict[str, Any], kind: str) -> tuple[str, dict[str, Any] | None]:
    if kind == "tb":
        item = choose_schedule(required_area, spandrel_long_items_for_dataset(dataset))
        return text((item or {}).get("mark")) or "—", item
    item = choose_schedule(required_area, spandrel_lig_items_for_dataset(dataset))
    return text((item or {}).get("mark")) or "—", item


def spandrel_quantities(item: dict[str, Any], dataset: dict[str, Any]) -> dict[str, float | None]:
    if item.get("kind") != "spandrel":
        return {"steel_kg": None, "concrete_volume_m3": None, "rate_kgpm3": None, "tb_kg": None, "ligs_kg": None}
    tb_item = item.get("current_tb_item") or spandrel_long_lookup_by_mark(item.get("current_tb_mark"), dataset)
    lig_item = item.get("current_lig_item") or spandrel_lig_lookup_by_mark(item.get("current_lig_mark"), dataset)
    H_mm = float(item.get("width_m") or 0.0) * 1000.0
    D_mm = float(item.get("depth_m") or 0.0) * 1000.0
    TW_mm = float(item.get("thickness_m") or 0.0) * 1000.0
    H_m = float(item.get("width_m") or 0.0)
    D_m = float(item.get("depth_m") or 0.0)
    TW_m = float(item.get("thickness_m") or 0.0)
    volume = D_m * TW_m * H_m if D_m > 0 and TW_m > 0 and H_m > 0 else None
    tb_bar = int((tb_item or {}).get("bar") or 0)
    lig_bar = int((lig_item or {}).get("bar") or 0)
    lig_legs = int((lig_item or {}).get("legs") or 0)
    ast_h_eff = float((tb_item or {}).get("area") or 0.0)
    ast_ligs_eff = float((lig_item or {}).get("area") or 0.0)
    lap_ratio_ligs = ((spandrel_cog_length_mm(lig_bar) * 2.0 * lig_legs) + D_mm) / D_mm if lig_bar and lig_legs and D_mm > 0 else None
    # Replicate workbook quantity logic exactly, including the use of the lig lap ratio in the top/bottom steel term.
    tb_kg = ast_h_eff * 2.0 * H_mm * float(lap_ratio_ligs or 0.0) * 7850.0 / 1_000_000_000.0 if ast_h_eff > 0 and H_mm > 0 and lap_ratio_ligs else None
    ligs_kg = ast_ligs_eff * float(lap_ratio_ligs or 0.0) * D_mm * H_mm * 7850.0 / 1_000_000_000_000.0 if ast_ligs_eff > 0 and D_mm > 0 and H_mm > 0 and lap_ratio_ligs else None
    steel_kg = sum(v for v in (tb_kg, ligs_kg) if v is not None) if any(v is not None for v in (tb_kg, ligs_kg)) else None
    rate = (steel_kg / volume) if (steel_kg is not None and volume and volume > 0) else None
    return {"steel_kg": steel_kg, "concrete_volume_m3": volume, "rate_kgpm3": rate, "tb_kg": tb_kg, "ligs_kg": ligs_kg}


def attach_design_and_overrides(dataset: dict[str, Any], item: dict[str, Any], include_boundary: bool = True) -> dict[str, Any]:
    result = deepcopy(item)
    design = deepcopy(dataset.get("design_by_key", {}).get(item["key"], {}))
    override = deepcopy(dataset.get("overrides", {}).get(item["key"], {}))
    fc = fc_for_material(dataset, item.get("material"))
    result["fc_mpa"] = fc
    result["story_height_m"] = float(dataset.get("height_by_story", {}).get(item.get("source_story")) or 0.0) or None

    if result["kind"] == "wall":
        zone_info = next((row for row in dataset.get("ductility_profile", {}).get("rows", []) if row.get("plot_story") == item.get("plot_story")), {})
        zone = text(zone_info.get("zone") or ("NO" if str((dataset.get("settings") or {}).get("ductility_level")) == "1" else "FULL"))
        min_pct = dataset.get("ductility_profile", {}).get("min_pct_by_plot_story", {}).get(item["plot_story"])
        min_vertical = (float(min_pct) / 100.0) * float(item.get("thickness_m") or 0.0) * 1_000_000.0 if min_pct is not None else None
        required_v = design.get("required_vertical_mm2pm")
        if min_vertical is not None:
            required_v = max(float(required_v or 0.0), float(min_vertical))
        required_h = design.get("required_horizontal_mm2pm")
        active_schedule = wall_schedule_items_for_dataset(dataset)
        seed_item_v = choose_schedule(required_v, active_schedule)
        seed_item_h = choose_schedule(required_h, active_schedule)
        seed_v = wall_schedule_text(seed_item_v)
        seed_h = wall_schedule_text(seed_item_h)

        current_v_text = current_reinf_value((override or {}).get("vertical"), seed_v)
        current_h_text = current_reinf_value((override or {}).get("horizontal"), seed_h)
        current_v_item = wall_schedule_lookup_by_reinf(current_v_text, dataset)
        current_h_item = wall_schedule_lookup_by_reinf(current_h_text, dataset)
        current_v_area = float((current_v_item or {}).get("area") or area_from_reinf(current_v_text) or 0.0) or None
        current_h_area = float((current_h_item or {}).get("area") or area_from_reinf(current_h_text) or 0.0) or None

        global_wall_type = text((dataset.get("settings") or {}).get("wall_type") or DEFAULT_WALL_TYPE)
        dowel_size = int((current_v_item or {}).get("dowel_size") or 0) if global_wall_type == "PRECAST" else 0
        dowel_length = (2.0 * wall_dowel_lap_length_mm(dowel_size) / 1000.0) * (1.5 if str((dataset.get("settings") or {}).get("ductility_level")) == "3" else 1.0) if dowel_size > 0 else 0.0
        actual_wall_type = "INSITU" if global_wall_type == "INSITU" or dowel_length <= 0.0 else "PRECAST"

        result.update({
            "required_vertical_mm2pm": required_v,
            "required_horizontal_mm2pm": required_h,
            "required_vertical_pct": design.get("required_vertical_pct"),
            "current_vertical_pct": design.get("current_vertical_pct"),
            "minimum_vertical_mm2pm": min_vertical,
            "minimum_vertical_pct": min_pct,
            "ductility_zone": zone,
            "suggested_vertical": seed_v,
            "suggested_horizontal": seed_h,
            "seed_vertical": seed_v,
            "seed_horizontal": seed_h,
            "override_vertical": text(override.get("vertical")),
            "override_horizontal": text(override.get("horizontal")),
            "current_vertical": current_v_text,
            "current_horizontal": current_h_text,
            "current_vertical_item": current_v_item,
            "current_horizontal_item": current_h_item,
            "current_vertical_mark": text((current_v_item or {}).get("mark")) or None,
            "current_horizontal_mark": text((current_h_item or {}).get("mark")) or None,
            "current_vertical_mm2pm": current_v_area,
            "current_horizontal_mm2pm": current_h_area,
            "display_vertical": current_v_text,
            "display_horizontal": current_h_text,
            "display_reinforcement": f"V: {current_v_text} / H: {current_h_text}",
            "wall_type": actual_wall_type,
            "is_overridden": bool(result.get("override_vertical") or result.get("override_horizontal")),
            "is_vertical_ok": required_v is None or (current_v_area or 0.0) >= float(required_v or 0.0) - 1e-9,
            "is_horizontal_ok": required_h is None or (current_h_area or 0.0) >= float(required_h or 0.0) - 1e-9,
            "messages": design.get("messages", []),
            "has_failure": bool(design.get("has_failure")),
        })
        if include_boundary:
            result.update(build_ligature_data(result, dataset))
        else:
            result.update({
                "ligs": "—", "ligs_n": None, "ligs_sets": None, "ligs_length_m": None,
                "has_boundary": False, "boundary_left_mm": 0.0, "boundary_right_mm": 0.0,
                "boundary_left_stress_ratio": None, "boundary_right_stress_ratio": None,
                "boundary_left_limit_ratio": None, "boundary_right_limit_ratio": None,
                "boundary_reg_mm": None, "duct3_bound_mm": None, "type1_ligs_required": None, "type2_ligs_required": None,
                "ligs_bar_mm": None, "ligs_spacing1_mm": None, "ligs_spacing2_mm": None, "extent_m": None,
                "horizontal_spacing_mm": None, "horizontal_spacing_warning": False,
            })
        result["body_hatch_style"] = body_hatch_style(fc) if include_boundary else None
        result["boundary_left_hatch_style"] = boundary_hatch_style(fc, result.get("boundary_left_stress_ratio")) if include_boundary and result.get("boundary_left_mm", 0) > 0 else None
        result["boundary_right_hatch_style"] = boundary_hatch_style(fc, result.get("boundary_right_stress_ratio")) if include_boundary and result.get("boundary_right_mm", 0) > 0 else None

        qty = wall_quantities(result, dataset)
        result.update({
            "steel_kg": qty["steel_kg"],
            "concrete_volume_m3": qty["concrete_volume_m3"],
            "current_rate_kgpm3": qty["rate_kgpm3"],
            "vertical_kg": qty["vertical_kg"],
            "horizontal_kg": qty["horizontal_kg"],
            "ligs_kg": qty["ligs_kg"],
            "dowels_kg": qty["dowels_kg"],
            "ubars_kg": qty["ubars_kg"],
            "vertical_bar_count_per_layer": qty["vertical_bar_count_per_layer"],
            "horizontal_bar_count_per_layer": qty["horizontal_bar_count_per_layer"],
            "dowel_size": qty["dowel_size"],
            "dowel_spacing": qty["dowel_spacing"],
            "no_dowels": qty["no_dowels"],
            "dowel_length_m": qty["dowel_length_m"],
            "ubar_size": qty["ubar_size"],
            "no_ubars": qty["no_ubars"],
            "ubar_length_m": qty["ubar_length_m"],
        })
        return result

    # spandrels
    required_tb = design.get("required_tb_area_mm2")
    ductility_level = str((dataset.get("settings") or {}).get("ductility_level") or DEFAULT_DUCTILITY_LEVEL)
    lig_factor = 1.0 if ductility_level == "1" else (2.0 / 0.77 if ductility_level == "2" else 3.0 / 0.67)
    required_lig = (float(design.get("required_lig_area_base_mm2pm") or 0.0) * lig_factor) if design.get("required_lig_area_base_mm2pm") is not None else None

    seed_tb_mark, seed_tb_item = spandrel_seed_from_design(required_tb, dataset, "tb")
    seed_lig_mark, seed_lig_item = spandrel_seed_from_design(required_lig, dataset, "lig")

    current_tb_mark = text((override or {}).get("top_bottom")) or seed_tb_mark
    current_lig_mark = text((override or {}).get("ligs")) or seed_lig_mark
    current_tb_item = spandrel_long_lookup_by_mark(current_tb_mark, dataset)
    current_lig_item = spandrel_lig_lookup_by_mark(current_lig_mark, dataset)

    current_tb_area = float((current_tb_item or {}).get("area") or 0.0) or None
    current_lig_area = float((current_lig_item or {}).get("area") or 0.0) or None

    result.update({
        "required_top_rebar_mm2": design.get("required_top_rebar_mm2"),
        "required_bottom_rebar_mm2": design.get("required_bottom_rebar_mm2"),
        "required_tb_area_mm2": required_tb,
        "required_lig_area_mm2pm": required_lig,
        "suggested_top_bottom": seed_tb_mark,
        "suggested_ligs": seed_lig_mark,
        "seed_top_bottom": seed_tb_mark,
        "seed_ligs": seed_lig_mark,
        "override_top_bottom": text((override or {}).get("top_bottom")),
        "override_ligs": text((override or {}).get("ligs")),
        "current_tb_mark": current_tb_mark,
        "current_lig_mark": current_lig_mark,
        "current_tb_item": current_tb_item,
        "current_lig_item": current_lig_item,
        "current_top_bottom": spandrel_long_display_text(current_tb_item),
        "current_ligs": spandrel_lig_display_text(current_lig_item),
        # provide aliases so the existing plot text path can still work
        "current_vertical": spandrel_long_display_text(current_tb_item),
        "current_horizontal": spandrel_lig_display_text(current_lig_item),
        "current_top_rebar_mm2": current_tb_area,
        "current_bottom_rebar_mm2": current_tb_area,
        "current_lig_area_mm2pm": current_lig_area,
        "display_reinforcement": f"T&B: {spandrel_long_display_text(current_tb_item)} / L: {spandrel_lig_display_text(current_lig_item)}",
        "is_overridden": bool(result.get("override_top_bottom") or result.get("override_ligs")),
        "is_tb_ok": required_tb is None or (current_tb_area or 0.0) >= float(required_tb or 0.0) - 1e-9,
        "is_lig_ok": required_lig is None or (current_lig_area or 0.0) >= float(required_lig or 0.0) - 1e-9,
        "messages": design.get("messages", []),
        "has_failure": bool(design.get("has_failure")),
    })
    qty = spandrel_quantities(result, dataset)
    result.update({
        "steel_kg": qty["steel_kg"],
        "concrete_volume_m3": qty["concrete_volume_m3"],
        "current_rate_kgpm3": qty["rate_kgpm3"],
        "tb_kg": qty["tb_kg"],
        "ligs_kg": qty["ligs_kg"],
    })
    return result


def build_wall_reinforcement_rows(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for item in items:
        if item["kind"] != "wall":
            continue
        rows.append({
            "key": item["key"], "label": item["label"], "story": item["plot_story"],
            "vertical": item.get("current_vertical") or "—", "horizontal": item.get("current_horizontal") or "—",
            "vertical_mark": item.get("current_vertical_mark"), "horizontal_mark": item.get("current_horizontal_mark"),
            "ligs": item.get("ligs") or "—", "ligs_n": item.get("ligs_n"), "ligs_length_m": item.get("ligs_length_m"),
            "vertical_area": item.get("current_vertical_mm2pm"), "horizontal_area": item.get("current_horizontal_mm2pm"),
            "rate_kgpm3": item.get("current_rate_kgpm3"),
            "steel_kg": item.get("steel_kg"), "concrete_volume_m3": item.get("concrete_volume_m3"),
            "vertical_kg": item.get("vertical_kg"), "horizontal_kg": item.get("horizontal_kg"), "ligs_kg": item.get("ligs_kg"),
            "dowels_kg": item.get("dowels_kg"), "ubars_kg": item.get("ubars_kg"),
            "vertical_bar_count_per_layer": item.get("vertical_bar_count_per_layer"), "horizontal_bar_count_per_layer": item.get("horizontal_bar_count_per_layer"),
            "required_vertical_area": item.get("required_vertical_mm2pm"), "required_horizontal_area": item.get("required_horizontal_mm2pm"),
            "vertical_ok": bool(item.get("is_vertical_ok")), "horizontal_ok": bool(item.get("is_horizontal_ok")), "overridden": bool(item.get("is_overridden")),
            "has_boundary": bool(item.get("has_boundary")), "horizontal_spacing_mm": item.get("horizontal_spacing_mm"), "horizontal_spacing_warning": bool(item.get("horizontal_spacing_warning")),
            "wall_type": item.get("wall_type"),
        })
    rows.sort(key=lambda r: (natural_sort_key(r["story"]), natural_sort_key(r["label"])))
    return rows


def build_spandrel_reinforcement_rows(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for item in items:
        if item["kind"] != "spandrel":
            continue
        rows.append({
            "key": item["key"], "label": item["label"], "story": item["plot_story"],
            "top_mark": item.get("current_tb_mark"), "bottom_mark": item.get("current_tb_mark"), "lig_mark": item.get("current_lig_mark"),
            "top_reo": item.get("current_top_bottom") or "—", "bottom_reo": item.get("current_top_bottom") or "—", "ligs": item.get("current_ligs") or "—",
            "required_top_mm2": item.get("required_top_rebar_mm2"), "required_bottom_mm2": item.get("required_bottom_rebar_mm2"), "required_ligs_mm2pm": item.get("required_lig_area_mm2pm"),
            "top_ok": bool(item.get("is_tb_ok")), "bottom_ok": bool(item.get("is_tb_ok")), "ligs_ok": bool(item.get("is_lig_ok")),
            "steel_kg": item.get("steel_kg"), "concrete_volume_m3": item.get("concrete_volume_m3"), "rate_kgpm3": item.get("current_rate_kgpm3"),
            "tb_kg": item.get("tb_kg"), "ligs_kg": item.get("ligs_kg"), "overridden": bool(item.get("is_overridden")),
        })
    rows.sort(key=lambda r: (natural_sort_key(r["story"]), natural_sort_key(r["label"])))
    return rows


def build_all_wall_reinforcement_rows(dataset: dict[str, Any], include_boundary: bool = True) -> list[dict[str, Any]]:
    items = [attach_design_and_overrides(dataset, item, include_boundary=include_boundary) for item in dataset.get("walls", [])]
    return build_wall_reinforcement_rows(items)


def build_all_spandrel_reinforcement_rows(dataset: dict[str, Any], include_boundary: bool = True) -> list[dict[str, Any]]:
    items = [attach_design_and_overrides(dataset, item, include_boundary=include_boundary) for item in dataset.get("spandrels", [])]
    return build_spandrel_reinforcement_rows(items)


def build_floor_summary_rows(dataset: dict[str, Any], include_boundary: bool = True) -> list[dict[str, Any]]:
    rows = []
    for plot_story in dataset.get("available_plot_stories", []):
        walls = [attach_design_and_overrides(dataset, w, include_boundary=include_boundary) for w in dataset.get("walls", []) if w["plot_story"] == plot_story]
        spandrels = [attach_design_and_overrides(dataset, s, include_boundary=include_boundary) for s in dataset.get("spandrels", []) if s["plot_story"] == plot_story]
        total_steel_kg = sum(float(i.get("steel_kg") or 0.0) for i in (walls + spandrels))
        total_volume_m3 = sum(float(i.get("concrete_volume_m3") or 0.0) for i in (walls + spandrels))
        rows.append({
            "plot_story": plot_story,
            "wall_count": len(walls),
            "spandrel_count": len(spandrels),
            "total_steel_kg": total_steel_kg,
            "total_volume_m3": total_volume_m3,
            "avg_rate_kgpm3": (total_steel_kg / total_volume_m3) if total_volume_m3 > 0 else None
        })
    return rows


def build_plan_payload(dataset: dict[str, Any], story_name: str) -> dict[str, Any]:
    include_boundary = truthy(request.args.get("show_boundary"))
    items = [attach_design_and_overrides(dataset, el, include_boundary=include_boundary) for el in (dataset["walls"] + dataset["spandrels"]) if el["plot_story"] == story_name]
    items.sort(key=lambda d: (d["kind"], natural_sort_key(d["label"])))
    return {
        "ok": True, "mode": "plan", "filename": dataset["filename"], "story": story_name,
        "summary": {"mode": "plan", "plot_story": story_name, "wall_count": sum(1 for i in items if i["kind"] == "wall"), "spandrel_count": sum(1 for i in items if i["kind"] == "spandrel")},
        "items": items, "has_boundary_workbook": bool(dataset.get("boundary_workbook"))
    }


def build_elevation_payload(dataset: dict[str, Any], grid_key: str) -> dict[str, Any]:
    grid = next((g for g in dataset.get("grids", []) if g["key"] == grid_key), None)
    if not grid:
        raise KeyError(f"Grid '{grid_key}' not found.")
    include_boundary = truthy(request.args.get("show_boundary"))
    items = []
    for base in dataset["walls"] + dataset["spandrels"]:
        if not element_intersects_grid(grid["axis"], float(grid["ordinate_m"]), base["corners"]):
            continue
        if alignment_to_grid(grid["axis"], float(base["angle_deg"])) < WALL_ELEVATION_ALIGNMENT_THRESHOLD:
            continue
        item = attach_design_and_overrides(dataset, base, include_boundary=include_boundary)
        station_m, display_width_m = projected_station_and_width(grid["axis"], item)
        if item["kind"] == "wall":
            z0_m = float(dataset["story_elevation_m"].get(item["plot_story"], 0.0))
            z1_m = z0_m + float(dataset["height_by_story"].get(item["source_story"]) or 0.0)
            x0_m = station_m - display_width_m / 2.0
            x1_m = station_m + display_width_m / 2.0
        else:
            if grid["axis"] == "X":
                x0_m = min(float(item.get("line_y1_m") or item["cgy_m"]), float(item.get("line_y2_m") or item["cgy_m"]))
                x1_m = max(float(item.get("line_y1_m") or item["cgy_m"]), float(item.get("line_y2_m") or item["cgy_m"]))
            else:
                x0_m = min(float(item.get("line_x1_m") or item["cgx_m"]), float(item.get("line_x2_m") or item["cgx_m"]))
                x1_m = max(float(item.get("line_x1_m") or item["cgx_m"]), float(item.get("line_x2_m") or item["cgx_m"]))
            if abs(x1_m - x0_m) < 0.03:
                x0_m -= float(item.get("thickness_m") or 0.0) / 2.0
                x1_m += float(item.get("thickness_m") or 0.0) / 2.0
            z1_m = float(dataset["story_elevation_m"].get(item["source_story"], 0.0))
            z0_m = z1_m - float(item.get("depth_m") or 0.0)

        item.update({"grid_key": grid_key, "grid_axis": grid["axis"], "grid_id": grid["id"], "grid_ordinate_m": float(grid["ordinate_m"]), "station_m": station_m, "display_width_m": display_width_m, "elev_x0_m": x0_m, "elev_x1_m": x1_m, "elev_z0_m": z0_m, "elev_z1_m": z1_m, "x0_m": x0_m, "x1_m": x1_m, "z0_m": z0_m, "z1_m": z1_m})
        items.append(item)

    items.sort(key=lambda d: (d.get("station_m", 0.0), natural_sort_key(d["label"])))
    return {
        "ok": True, "mode": "elevation", "filename": dataset["filename"], "grid": grid, "story_elevation_m": dataset["story_elevation_m"], "stories_top_down": dataset["stories_top_down"], "items": items,
        "summary": {"mode": "elevation", "grid_id": grid["id"], "wall_count": sum(1 for i in items if i["kind"] == "wall"), "spandrel_count": sum(1 for i in items if i["kind"] == "spandrel")},
        "has_boundary_workbook": bool(dataset.get("boundary_workbook"))
    }


def remove_boundary_dataset(boundary_info: dict[str, Any] | None) -> None:
    if not boundary_info:
        return
    path = boundary_info.get("path")
    if path:
        Path(path).unlink(missing_ok=True)
    boundary_id = text(boundary_info.get("id"))
    if boundary_id:
        BOUNDARY_DATASETS.pop(boundary_id, None)


def clear_boundary_from_dataset(dataset: dict[str, Any]) -> None:
    remove_boundary_dataset(dataset.get("boundary_workbook"))
    dataset["boundary_workbook"] = None
    dataset["boundary_by_key"] = {}


def remove_dataset(dataset_id: str) -> None:
    dataset = DATASETS.pop(dataset_id, None)
    if not dataset:
        return
    clear_boundary_from_dataset(dataset)
    path = dataset.get("path")
    if path:
        Path(path).unlink(missing_ok=True)


@app.errorhandler(404)
def handle_404(err):
    if request.path.startswith("/api/"):
        return jsonify({"ok": False, "error": f"API route not found: {request.path}"}), 404
    return err


@app.errorhandler(405)
def handle_405(err):
    if request.path.startswith("/api/"):
        return jsonify({"ok": False, "error": f"Method not allowed for API route: {request.path}"}), 405
    return err


@app.errorhandler(500)
def handle_500(err):
    if request.path.startswith("/api/"):
        return jsonify({"ok": False, "error": "Internal server error."}), 500
    return err


@app.get("/")
def root():
    return send_file(FRONTEND_FILE)


def upload_response_payload(dataset_id: str, dataset: dict[str, Any]) -> dict[str, Any]:
    return {
        "ok": True,
        "dataset_id": dataset_id,
        "filename": dataset["filename"],
        "stories": dataset["available_plot_stories"],
        "default_story": dataset["available_plot_stories"][0] if dataset["available_plot_stories"] else None,
        "grids": dataset["grids"],
        "default_grid_key": dataset["default_grid_key"],
        "has_pier_design": dataset["has_pier_design"],
        "has_spandrels": dataset["has_spandrels"],
        "wall_reinforcement_schedule": wall_schedule_items_for_dataset(dataset, active_only=False),
        "reinforcement_schedule": wall_schedule_items_for_dataset(dataset, active_only=False),
        "spandrel_tb_schedule": spandrel_long_items_for_dataset(dataset, active_only=False),
        "spandrel_lig_schedule": spandrel_lig_items_for_dataset(dataset, active_only=False),
        "schedule_active_marks": {
            "wall": dataset.get("wall_schedule_active_marks", []),
            "spandrel_tb": dataset.get("spandrel_tb_active_marks", []),
            "spandrel_lig": dataset.get("spandrel_lig_active_marks", []),
        },
        "active_schedule_marks": dataset.get("wall_schedule_active_marks", []),
        "ductility_profile": dataset.get("ductility_profile"),
        "grouping_story_list": dataset.get("grouping_story_list", []) or dataset.get("available_plot_stories", []),
        "has_boundary_workbook": bool(dataset.get("boundary_workbook")),
        "settings": dataset.get("settings", {}),
        "pier_failure_count": dataset.get("pier_failure_count", 0),
        "spandrel_failure_count": dataset.get("spandrel_failure_count", 0),
    }


@app.post("/api/upload")
def upload_workbook():
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "No file uploaded."}), 400
    file = request.files["file"]
    if not file or not file.filename:
        return jsonify({"ok": False, "error": "No file selected."}), 400
    filename = secure_filename(file.filename)
    ext = Path(filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({"ok": False, "error": "Please upload an Excel workbook (.xlsx, .xlsm, .xltx, .xltm)."}), 400
    previous_dataset_id = text(request.form.get("previous_dataset_id"))
    dataset_id = uuid.uuid4().hex[:12]
    save_path = UPLOAD_DIR / f"{dataset_id}_{filename}"
    file.save(save_path)
    try:
        parsed = parse_workbook(save_path)
    except Exception as exc:
        save_path.unlink(missing_ok=True)
        return jsonify({"ok": False, "error": f"Could not read workbook: {exc}"}), 400
    parsed["path"] = str(save_path)
    DATASETS[dataset_id] = parsed
    if previous_dataset_id and previous_dataset_id != dataset_id:
        try:
            remove_dataset(previous_dataset_id)
        except Exception:
            pass
    return jsonify(upload_response_payload(dataset_id, parsed))


@app.post("/api/upload-boundary")
def upload_boundary_workbook():
    dataset_id = text(request.form.get("dataset_id"))
    dataset = DATASETS.get(dataset_id)
    if not dataset:
        return jsonify({"ok": False, "error": "Main workbook dataset not found."}), 404
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "No file uploaded."}), 400
    file = request.files["file"]
    if not file or not file.filename:
        return jsonify({"ok": False, "error": "No file selected."}), 400
    filename = secure_filename(file.filename)
    ext = Path(filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({"ok": False, "error": "Please upload an Excel workbook (.xlsx, .xlsm, .xltx, .xltm)."}), 400
    clear_boundary_from_dataset(dataset)
    boundary_id = uuid.uuid4().hex[:12]
    save_path = UPLOAD_DIR / f"boundary_{dataset_id}_{boundary_id}_{filename}"
    file.save(save_path)
    try:
        boundary_by_key = parse_boundary_workbook(save_path)
    except Exception as exc:
        save_path.unlink(missing_ok=True)
        return jsonify({"ok": False, "error": f"Could not read boundary workbook: {exc}"}), 400
    boundary_info = {"id": boundary_id, "filename": filename, "path": str(save_path)}
    BOUNDARY_DATASETS[boundary_id] = boundary_info
    dataset["boundary_workbook"] = boundary_info
    dataset["boundary_by_key"] = boundary_by_key
    return jsonify({"ok": True, "dataset_id": dataset_id, "boundary_dataset_id": boundary_id, "filename": filename, "boundary_count": len(boundary_by_key)})


@app.post("/api/dataset/<dataset_id>/boundary/clear")
def clear_boundary_workbook(dataset_id: str):
    dataset = DATASETS.get(dataset_id)
    if not dataset:
        return jsonify({"ok": False, "error": "Dataset not found."}), 404
    clear_boundary_from_dataset(dataset)
    return jsonify({"ok": True})


@app.post("/api/dataset/<dataset_id>/settings")
def set_settings(dataset_id: str):
    dataset = DATASETS.get(dataset_id)
    if not dataset:
        return jsonify({"ok": False, "error": "Dataset not found."}), 404
    payload = request.get_json(silent=True) or {}
    settings = deepcopy(dataset.get("settings") or {})
    level = text(payload.get("level") or payload.get("ductility_level") or settings.get("ductility_level") or DEFAULT_DUCTILITY_LEVEL) or DEFAULT_DUCTILITY_LEVEL
    wall_type = text(payload.get("wall_type") or settings.get("wall_type") or DEFAULT_WALL_TYPE).upper()
    if wall_type not in {"INSITU", "PRECAST"}:
        wall_type = DEFAULT_WALL_TYPE
    cover_mm = as_float(payload.get("cover_mm"), as_float(settings.get("cover_mm"), DEFAULT_COVER_MM))
    settings.update({"ductility_level": level, "wall_type": wall_type, "cover_mm": float(cover_mm or DEFAULT_COVER_MM)})
    dataset["settings"] = settings
    dataset["ductility_profile"] = build_ductility_profile(dataset, level)
    return jsonify({"ok": True, "settings": settings, "ductility_profile": dataset["ductility_profile"], "grouping_story_list": dataset.get("grouping_story_list", []) or dataset.get("available_plot_stories", [])})


@app.post("/api/dataset/<dataset_id>/ductility")
def set_ductility(dataset_id: str):
    payload = request.get_json(silent=True) or {}
    payload["level"] = payload.get("level")
    return set_settings(dataset_id)


@app.post("/api/dataset/<dataset_id>/schedule-marks")
def set_schedule_marks(dataset_id: str):
    dataset = DATASETS.get(dataset_id)
    if not dataset:
        return jsonify({"ok": False, "error": "Dataset not found."}), 404
    payload = request.get_json(silent=True) or {}
    if isinstance(payload.get("active_marks"), list):
        requested_wall = payload.get("active_marks")
        requested_sp_tb = dataset.get("spandrel_tb_active_marks", [])
        requested_sp_lig = dataset.get("spandrel_lig_active_marks", [])
    else:
        requested_wall = payload.get("wall_active_marks")
        requested_sp_tb = payload.get("spandrel_tb_active_marks")
        requested_sp_lig = payload.get("spandrel_lig_active_marks")
    if not isinstance(requested_wall, list) or not isinstance(requested_sp_tb, list) or not isinstance(requested_sp_lig, list):
        return jsonify({"ok": False, "error": "Schedule marks payload is invalid."}), 400
    valid_wall = {text(item.get("mark")) for item in DEFAULT_WALL_REINF_SCHEDULE}
    valid_sp_tb = {text(item.get("mark")) for item in DEFAULT_SPANDREL_LONG_SCHEDULE}
    valid_sp_lig = {text(item.get("mark")) for item in DEFAULT_SPANDREL_LIG_SCHEDULE}
    dataset["wall_schedule_active_marks"] = [text(item.get("mark")) for item in DEFAULT_WALL_REINF_SCHEDULE if text(item.get("mark")) in {text(v) for v in requested_wall if text(v) in valid_wall}]
    dataset["spandrel_tb_active_marks"] = [text(item.get("mark")) for item in DEFAULT_SPANDREL_LONG_SCHEDULE if text(item.get("mark")) in {text(v) for v in requested_sp_tb if text(v) in valid_sp_tb}]
    dataset["spandrel_lig_active_marks"] = [text(item.get("mark")) for item in DEFAULT_SPANDREL_LIG_SCHEDULE if text(item.get("mark")) in {text(v) for v in requested_sp_lig if text(v) in valid_sp_lig}]
    return jsonify({
        "ok": True,
        "active_marks": {
            "wall": dataset.get("wall_schedule_active_marks", []),
            "spandrel_tb": dataset.get("spandrel_tb_active_marks", []),
            "spandrel_lig": dataset.get("spandrel_lig_active_marks", []),
        },
        "active_schedule_marks": dataset.get("wall_schedule_active_marks", []),
        "wall_reinforcement_schedule": wall_schedule_items_for_dataset(dataset, active_only=False),
        "reinforcement_schedule": wall_schedule_items_for_dataset(dataset, active_only=False),
        "spandrel_tb_schedule": spandrel_long_items_for_dataset(dataset, active_only=False),
        "spandrel_lig_schedule": spandrel_lig_items_for_dataset(dataset, active_only=False),
    })


@app.get("/api/dataset/<dataset_id>/plan/<story_name>")
def get_plan(dataset_id: str, story_name: str):
    dataset = DATASETS.get(dataset_id)
    if not dataset:
        return jsonify({"ok": False, "error": "Dataset not found."}), 404
    if story_name not in dataset["available_plot_stories"]:
        return jsonify({"ok": False, "error": f"Story '{story_name}' not found."}), 404
    return jsonify(build_plan_payload(dataset, story_name))


@app.get("/api/dataset/<dataset_id>/elevation/<path:grid_key>")
def get_elevation(dataset_id: str, grid_key: str):
    dataset = DATASETS.get(dataset_id)
    if not dataset:
        return jsonify({"ok": False, "error": "Dataset not found."}), 404
    try:
        return jsonify(build_elevation_payload(dataset, grid_key))
    except KeyError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 404


@app.get("/api/dataset/<dataset_id>/reinforcement")
def get_reinforcement(dataset_id: str):
    dataset = DATASETS.get(dataset_id)
    if not dataset:
        return jsonify({"ok": False, "error": "Dataset not found."}), 404
    include_boundary = truthy(request.args.get("use_boundary"))
    return jsonify({"ok": True, "rows": build_all_wall_reinforcement_rows(dataset, include_boundary=include_boundary)})


@app.get("/api/dataset/<dataset_id>/spandrels-reinforcement")
def get_spandrels_reinforcement(dataset_id: str):
    dataset = DATASETS.get(dataset_id)
    if not dataset:
        return jsonify({"ok": False, "error": "Dataset not found."}), 404
    include_boundary = truthy(request.args.get("use_boundary"))
    return jsonify({"ok": True, "rows": build_all_spandrel_reinforcement_rows(dataset, include_boundary=include_boundary)})


@app.get("/api/dataset/<dataset_id>/summary")
def get_summary(dataset_id: str):
    dataset = DATASETS.get(dataset_id)
    if not dataset:
        return jsonify({"ok": False, "error": "Dataset not found."}), 404
    include_boundary = truthy(request.args.get("use_boundary"))
    return jsonify({
        "ok": True,
        "rows": build_floor_summary_rows(dataset, include_boundary=include_boundary),
        "ductility_profile": dataset.get("ductility_profile"),
        "grouping_story_list": dataset.get("grouping_story_list", []) or dataset.get("available_plot_stories", []),
        "pier_failure_count": dataset.get("pier_failure_count", 0),
        "spandrel_failure_count": dataset.get("spandrel_failure_count", 0),
    })


@app.get("/api/dataset/<dataset_id>/item/<path:item_key>")
def get_item(dataset_id: str, item_key: str):
    dataset = DATASETS.get(dataset_id)
    if not dataset:
        return jsonify({"ok": False, "error": "Dataset not found."}), 404
    base = next((el for el in (dataset["walls"] + dataset["spandrels"]) if el["key"] == item_key), None)
    if not base:
        return jsonify({"ok": False, "error": "Element not found."}), 404
    include_boundary = truthy(request.args.get("use_boundary")) if request.args.get("use_boundary") is not None else bool(dataset.get("boundary_workbook"))
    return jsonify({"ok": True, "item": attach_design_and_overrides(dataset, base, include_boundary=include_boundary)})


@app.post("/api/dataset/<dataset_id>/override")
def set_override(dataset_id: str):
    dataset = DATASETS.get(dataset_id)
    if not dataset:
        return jsonify({"ok": False, "error": "Dataset not found."}), 404
    payload = request.get_json(silent=True) or {}
    key = text(payload.get("key"))
    if not key:
        return jsonify({"ok": False, "error": "Missing element key."}), 400
    base = next((el for el in (dataset["walls"] + dataset["spandrels"]) if el["key"] == key), None)
    if not base:
        return jsonify({"ok": False, "error": "Element not found."}), 404

    if key.startswith("wall|"):
        vertical = text(payload.get("vertical"))
        horizontal = text(payload.get("horizontal"))
        if not vertical and not horizontal:
            dataset["overrides"].pop(key, None)
        else:
            dataset["overrides"][key] = {"vertical": vertical, "horizontal": horizontal}
    else:
        top_bottom = text(payload.get("top_bottom"))
        ligs = text(payload.get("ligs"))
        if not top_bottom and not ligs:
            dataset["overrides"].pop(key, None)
        else:
            dataset["overrides"][key] = {"top_bottom": top_bottom, "ligs": ligs}

    include_boundary = truthy(payload.get("use_boundary")) if payload.get("use_boundary") is not None else bool(dataset.get("boundary_workbook"))
    item = attach_design_and_overrides(dataset, base, include_boundary=include_boundary)
    warnings = []
    if item.get("kind") == "wall":
        if not item.get("is_vertical_ok", True):
            warnings.append("Vertical override area is below the required minimum.")
        if not item.get("is_horizontal_ok", True):
            warnings.append("Horizontal override area is below the required minimum.")
    else:
        if not item.get("is_tb_ok", True):
            warnings.append("Top/bottom override area is below the required minimum.")
        if not item.get("is_lig_ok", True):
            warnings.append("Ligature override area is below the required minimum.")
    return jsonify({"ok": True, "item": item, "warning": " ".join(warnings)})


@app.post("/api/dataset/<dataset_id>/override/reset-all")
def reset_all_overrides(dataset_id: str):
    dataset = DATASETS.get(dataset_id)
    if not dataset:
        return jsonify({"ok": False, "error": "Dataset not found."}), 404
    dataset["overrides"] = {}
    return jsonify({"ok": True})


@app.get("/api/health")
def health():
    return jsonify({"ok": True})


if __name__ == "__main__":
    app.run(debug=False, use_reloader=False, host="127.0.0.1", port=5000)
